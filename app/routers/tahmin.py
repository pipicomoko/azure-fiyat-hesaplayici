"""Tahmin calisma alani: urun ekleme, alan degisiminde aninda yeniden
hesaplama, kaldirma, kaydetme, Excel'e aktarma ve dil degistirme.

Tum kalemler TEK bir <form id="tahmin-formu"> icinde yasar (bkz.
templates/tahmin.html); her kalemin alanlari `{kalem_id}.{alan}` seklinde
adlandirilir (bkz. app/form_yardimcilari.py). htmx, bir forma ait elemanlarda
varsayilan olarak en yakin formun TUM degerlerini istekle birlikte gonderir;
bu sayede tek bir kalem degisince o kalemi (URL'deki ?kalem_id ile) izole
edip yeniden hesaplayabiliriz, ayni zamanda dil degisimi/kaydetme/disa
aktarim gibi TOPLU islemler de ayni formdan TUM kalemleri toplu okuyabilir.

Fiyat, HER ZAMAN bu istekte yeniden hesaplanir (asla istemciden gelen bir
fiyat degerine guvenilmez) -- boylece "no fabricated price" kurali ve
"Excel'deki rakamlar ekranla birebir tutarli olmali" kurali ayni anda,
tek bir gercek kaynaktan saglanir.
"""

from __future__ import annotations

import io
import uuid
from dataclasses import dataclass
from datetime import datetime

from fastapi import APIRouter, Depends, Request
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from sqlmodel import Session, select

from app.database import oturum_al
from app.disa_aktar import TahminBosHatasi, calisma_kitabi_olustur
from app.fiyat_api import FiyatApiHatasi
from app.form_yardimcilari import (
    bos_degerleri_temizle,
    boolean_alanlarini_normallestir,
    coklu_kalem_formunu_ayir,
)
from app.i18n import DESTEKLENEN_DILLER, DIL_COOKIE_ADI, Dil, form_alanindan_dil_al, istekten_dil_al, t
from app.kayitli_tahmin import hesaplamadan_satirlar
from app.models import Hesaplama, HesaplamaKalemi
from app.para_birimleri import PARA_BIRIMLERI, VARSAYILAN_PARA_BIRIMI, guvenli_para_birimi
from app.products import KAYITLI_URUNLER, urun_al
from app.products.base import DisaAktarimSatiri, FiyatBulunamadiHatasi, FiyatSonucu, UrunModulu
from app.sablonlar import render, templates
from app.yetkilendirme import (
    IZIN_HESAPLAMA_KULLAN,
    departman_etiketi,
    gecmis_erisim_kapsami,
    gecmis_goruntule_gerekli,
    gruplardan_departman_belirle,
    hesaplama_departmani,
    hesaplama_gorunen_durum,
    hesaplamaya_erisebilir_mi,
    hesaplamayi_duzenleyebilir_mi,
    hesaplamayi_silebilir_mi,
    yetki_gerekli,
)

router = APIRouter(prefix="/tahmin")


@dataclass
class _KalemSonucu:
    kalem_id: str
    urun_tipi: str
    urun: UrunModulu
    yapilandirma: dict
    secenekler: dict
    gorunur_alanlar: set
    fiyat: FiyatSonucu | None
    hata: str | None


async def _fiyatla_guvenli(urun: UrunModulu, yapilandirma: dict, para_birimi: str, dil: Dil):
    try:
        return await urun.fiyatla(yapilandirma, para_birimi), None
    except FiyatBulunamadiHatasi:
        return None, t("fiyat_bulunamadi", dil)
    except FiyatApiHatasi:
        return None, t("fiyat_servisi_erisilemez", dil)
    except (TypeError, ValueError):
        return None, t("fiyat_bulunamadi", dil)


async def _kalemi_coz(kalem_id: str, ham_yapilandirma: dict, para_birimi: str, dil: Dil) -> _KalemSonucu | None:
    ham = dict(ham_yapilandirma)
    bos_degerleri_temizle(ham)
    boolean_alanlarini_normallestir(ham)
    urun_tipi = ham.pop("urun_tipi", "")
    urun = urun_al(urun_tipi)
    if urun is None:
        return None

    # Secenek cozumleme (ozellikle VM katalogu) API'ye baglidir; FiyatApiHatasi
    # burada yukselir — kalem-ekle/hesapla yakalayip kullaniciya uyari gosterir.
    secenek_sonucu = await urun.secenekleri_getir(ham, dil)
    yapilandirma = secenek_sonucu.yapilandirma
    fiyat, hata = await _fiyatla_guvenli(urun, yapilandirma, para_birimi, dil)
    return _KalemSonucu(
        kalem_id, urun_tipi, urun, yapilandirma, secenek_sonucu.secenekler,
        secenek_sonucu.gorunur_alanlar, fiyat, hata,
    )


def _fiyat_servisi_uyarisi(dil: Dil) -> HTMLResponse:
    """kalem-ekle / kalem-hesapla basarisizliginda kullaniciya gorunur uyari."""
    metin = t("fiyat_servisi_erisilemez", dil)
    html = f'<div class="ui-alert ui-alert--warning" role="alert" style="margin: var(--space-3)">{metin}</div>'
    return HTMLResponse(html, status_code=200)



async def _tum_kalemleri_coz(kalemler_ham: dict[str, dict], para_birimi: str, dil: Dil) -> list[_KalemSonucu]:
    sonuclar = []
    for kalem_id, ham in kalemler_ham.items():
        sonuc = await _kalemi_coz(kalem_id, ham, para_birimi, dil)
        if sonuc is not None:
            sonuclar.append(sonuc)
    return sonuclar


def _kalem_baglami(kalem_id: str, sonuc: _KalemSonucu, para_birimi: str, dil: Dil) -> dict:
    return {
        "dil": dil,
        "urun": sonuc.urun,
        "urun_tipi": sonuc.urun_tipi,
        "kalem_id": kalem_id,
        "yapilandirma": sonuc.yapilandirma,
        "secenekler": sonuc.secenekler,
        "gorunur_alanlar": sonuc.gorunur_alanlar,
        "para_birimi": para_birimi,
        "fiyat": sonuc.fiyat,
        "hata": sonuc.hata,
    }


@router.get("")
async def tahmin_sayfasi(
    request: Request,
    oturum: Session = Depends(oturum_al),
    kullanici: dict = Depends(yetki_gerekli(IZIN_HESAPLAMA_KULLAN)),
):
    dil = istekten_dil_al(request)
    para_birimi = VARSAYILAN_PARA_BIRIMI
    kalem_sonuclari: list[tuple[str, _KalemSonucu]] = []
    hesaplama_adi = ""
    duzenlenen_id: int | None = None
    red_gerekce = None

    ham_id = request.query_params.get("hesaplama_id")
    if ham_id:
        try:
            hesaplama_id = int(ham_id)
        except ValueError:
            hesaplama_id = None
        if hesaplama_id is not None:
            hesaplama = oturum.get(Hesaplama, hesaplama_id)
            if hesaplama is None or not hesaplamayi_duzenleyebilir_mi(kullanici, hesaplama):
                return HTMLResponse(t("gecmis_yalnizca_sahibi_erisebilir", dil), status_code=403)
            duzenlenen_id = hesaplama.id
            hesaplama_adi = hesaplama.ad or ""
            para_birimi = guvenli_para_birimi(hesaplama.para_birimi)
            red_gerekce = hesaplama.red_gerekce
            for kayitli in hesaplama.kalemler or []:
                kalem_id = uuid.uuid4().hex
                ham = dict(kayitli.yapilandirma or {})
                ham["urun_tipi"] = kayitli.urun_tipi
                if kayitli.indirim_yuzdesi is not None:
                    ham["indirim_yuzdesi"] = kayitli.indirim_yuzdesi
                sonuc = await _kalemi_coz(kalem_id, ham, para_birimi, dil)
                if sonuc is None:
                    continue
                if kayitli.indirim_yuzdesi is not None:
                    sonuc.yapilandirma = {
                        **sonuc.yapilandirma,
                        "indirim_yuzdesi": kayitli.indirim_yuzdesi,
                    }
                kalem_sonuclari.append((kalem_id, sonuc))

    return render(
        request,
        "tahmin.html",
        {
            "urunler": list(KAYITLI_URUNLER.values()),
            "para_birimleri": PARA_BIRIMLERI,
            "para_birimi": para_birimi,
            "tahmin_modu": True,
            "kalem_sonuclari": kalem_sonuclari,
            "hesaplama_adi": hesaplama_adi,
            "duzenlenen_hesaplama_id": duzenlenen_id,
            "red_gerekce": red_gerekce,
        },
    )


@router.post("/kalem-ekle")
async def kalem_ekle(request: Request, kullanici: dict = Depends(yetki_gerekli(IZIN_HESAPLAMA_KULLAN))):
    form = await request.form()
    urun_tipi = form.get("urun_tipi", "")
    para_birimi = guvenli_para_birimi(form.get("para_birimi"))
    dil = istekten_dil_al(request)

    urun = urun_al(urun_tipi)
    if urun is None:
        return HTMLResponse("", status_code=400)

    kalem_id = uuid.uuid4().hex
    try:
        sonuc = await _kalemi_coz(kalem_id, {**urun.bos_yapilandirma(), "urun_tipi": urun_tipi}, para_birimi, dil)
    except FiyatApiHatasi:
        return _fiyat_servisi_uyarisi(dil)
    if sonuc is None:
        return HTMLResponse("", status_code=400)

    return templates.TemplateResponse(request, urun.sablon_adi, _kalem_baglami(kalem_id, sonuc, para_birimi, dil))


@router.post("/kalem/hesapla")
async def kalem_hesapla(request: Request, kullanici: dict = Depends(yetki_gerekli(IZIN_HESAPLAMA_KULLAN))):
    kalem_id = request.query_params.get("kalem_id", "")
    form = await request.form()
    genel, kalemler = coklu_kalem_formunu_ayir(form)

    ham = kalemler.get(kalem_id)
    if ham is None:
        return HTMLResponse("", status_code=400)

    para_birimi = guvenli_para_birimi(genel.get("para_birimi"))
    dil = form_alanindan_dil_al(genel.get("dil")) if genel.get("dil") in DESTEKLENEN_DILLER else istekten_dil_al(request)

    try:
        sonuc = await _kalemi_coz(kalem_id, ham, para_birimi, dil)
    except FiyatApiHatasi:
        return _fiyat_servisi_uyarisi(dil)
    if sonuc is None:
        return HTMLResponse("", status_code=400)

    return templates.TemplateResponse(request, sonuc.urun.sablon_adi, _kalem_baglami(kalem_id, sonuc, para_birimi, dil))


@router.post("/dil-degistir")
async def dil_degistir_tahmin(request: Request, kullanici: dict = Depends(yetki_gerekli(IZIN_HESAPLAMA_KULLAN))):
    form = await request.form()
    genel, kalemler_ham = coklu_kalem_formunu_ayir(form)

    yeni_dil = form_alanindan_dil_al(genel.get("dil"))
    para_birimi = guvenli_para_birimi(genel.get("para_birimi"))

    kalem_sonuclari = await _tum_kalemleri_coz(kalemler_ham, para_birimi, yeni_dil)

    baglam = {
        "dil": yeni_dil,
        "kullanici": request.session.get("kullanici"),
        "urunler": list(KAYITLI_URUNLER.values()),
        "para_birimleri": PARA_BIRIMLERI,
        "para_birimi": para_birimi,
        "kalem_sonuclari": [(k.kalem_id, k) for k in kalem_sonuclari],
        "tahmin_modu": True,
    }
    # Ana icerigi VE nav'i (dil etiketleri icin out-of-band swap olarak) ayni
    # yanitta doner -- boylece tek istekte hem tahmin hem de sayfa govdesindeki
    # sabit metinler yeni dile gecer, tahmin ASLA kaybolmaz.
    icerik_html = templates.env.get_template("_tahmin_ic_icerik.html").render(request=request, **baglam)
    nav_html = templates.env.get_template("_nav.html").render(request=request, oob=True, **baglam)

    yanit = HTMLResponse(icerik_html + nav_html)
    yanit.set_cookie(DIL_COOKIE_ADI, yeni_dil, max_age=60 * 60 * 24 * 365, samesite="lax")
    return yanit


@router.post("/kaydet")
async def tahmin_kaydet(
    request: Request,
    oturum: Session = Depends(oturum_al),
    kullanici: dict = Depends(yetki_gerekli(IZIN_HESAPLAMA_KULLAN)),
):
    dil = istekten_dil_al(request)
    form = await request.form()
    genel, kalemler_ham = coklu_kalem_formunu_ayir(form)
    para_birimi = guvenli_para_birimi(genel.get("para_birimi"))
    hesaplama_adi = (genel.get("hesaplama_adi") or "").strip()

    if not hesaplama_adi:
        return HTMLResponse(f'<div class="ui-alert ui-alert--danger">{t("kaydet_ad_gerekli", dil)}</div>', status_code=400)
    if not kalemler_ham:
        return HTMLResponse(f'<div class="ui-alert ui-alert--danger">{t("kaydet_bos_sepet", dil)}</div>', status_code=400)

    kalem_sonuclari = await _tum_kalemleri_coz(kalemler_ham, para_birimi, dil)
    if not kalem_sonuclari or any(k.hata for k in kalem_sonuclari):
        return HTMLResponse(f'<div class="ui-alert ui-alert--danger">{t("fiyat_bulunamadi", dil)}</div>', status_code=400)

    onaya_gonder = (genel.get("onaya_gonder") or "").lower() in {"1", "true", "evet", "on"}
    onay_hedefi = (genel.get("onay_hedefi") or "").strip().lower() or None

    from app.models import DURUM_ONAY_BEKLIYOR, DURUM_TASLAK
    from app.yetkilendirme import (
        oturum_manager_adi,
        oturum_manager_zincirini_genislet,
        ustu_olmayan_mi,
    )

    zincir = oturum_manager_zincirini_genislet(kullanici)
    # Ustunde kimse yoksa onay akisina hicbir sekilde giremez
    if ustu_olmayan_mi(kullanici):
        onaya_gonder = False
        onay_hedefi = None

    if onaya_gonder and not onay_hedefi:
        return HTMLResponse(
            f'<div class="ui-alert ui-alert--danger">{t("onay_hedefi_gerekli", dil)}</div>',
            status_code=400,
        )

    departman_anahtari = kullanici.get("departman")
    if (not departman_anahtari or departman_anahtari == "diger") and kullanici.get("unvan"):
        from app.yetkilendirme import _departman_anahtari_etiketten_turetilir as _dep_turet
        turetilen = _dep_turet(kullanici["unvan"])
        if turetilen and turetilen != "diger":
            departman_anahtari = turetilen
    if not departman_anahtari or departman_anahtari == "diger":
        departman_anahtari, _ = gruplardan_departman_belirle(kullanici.get("gruplar"))

    if onaya_gonder and onay_hedefi and onay_hedefi not in zincir:
        zincir = [onay_hedefi] + [z for z in zincir if z != onay_hedefi]
    onay_hedefi_ad = oturum_manager_adi(kullanici, onay_hedefi) if onay_hedefi else None

    durum = DURUM_ONAY_BEKLIYOR if (onaya_gonder and onay_hedefi) else DURUM_TASLAK

    duzenlenen_id = None
    ham_id = (genel.get("hesaplama_id") or "").strip()
    if ham_id:
        try:
            duzenlenen_id = int(ham_id)
        except ValueError:
            duzenlenen_id = None

    hesaplama: Hesaplama | None = None
    if duzenlenen_id is not None:
        hesaplama = oturum.get(Hesaplama, duzenlenen_id)
        if hesaplama is None or not hesaplamayi_duzenleyebilir_mi(kullanici, hesaplama):
            return HTMLResponse(
                f'<div class="alert alert-danger">{t("gecmis_yalnizca_sahibi_erisebilir", dil)}</div>',
                status_code=403,
            )
        # Eski kalemleri sil, ust kaydi guncelle
        for eski in list(hesaplama.kalemler or []):
            oturum.delete(eski)
        hesaplama.ad = hesaplama_adi
        hesaplama.para_birimi = para_birimi
        hesaplama.olusturan_gruplar = list(kullanici.get("gruplar", []))
        hesaplama.olusturan_departman = departman_anahtari
        hesaplama.olusturan_unvan = kullanici.get("unvan") or ""
        hesaplama.olusturan_ad_soyad = kullanici.get("ad_soyad") or kullanici.get("kullanici_adi") or ""
        hesaplama.durum = durum
        hesaplama.onay_hedefi = onay_hedefi if durum == DURUM_ONAY_BEKLIYOR else None
        hesaplama.onay_hedefi_ad_soyad = onay_hedefi_ad if durum == DURUM_ONAY_BEKLIYOR else None
        hesaplama.olusturan_manager_zinciri = zincir
        hesaplama.red_gerekce = None
        if durum == DURUM_ONAY_BEKLIYOR:
            hesaplama.onaylayan_kullanici_adi = None
            hesaplama.onay_tarihi = None
    else:
        hesaplama = Hesaplama(
            ad=hesaplama_adi,
            para_birimi=para_birimi,
            toplam_aylik_maliyet=0.0,
            olusturan_kullanici_adi=kullanici["kullanici_adi"],
            olusturan_gruplar=list(kullanici.get("gruplar", [])),
            olusturan_departman=departman_anahtari,
            olusturan_unvan=kullanici.get("unvan") or "",
            olusturan_ad_soyad=kullanici.get("ad_soyad") or kullanici.get("kullanici_adi") or "",
            durum=durum,
            revizyon=1,
            onay_hedefi=onay_hedefi if durum == DURUM_ONAY_BEKLIYOR else None,
            onay_hedefi_ad_soyad=onay_hedefi_ad if durum == DURUM_ONAY_BEKLIYOR else None,
            olusturan_manager_zinciri=zincir,
        )
        oturum.add(hesaplama)
        oturum.flush()

    toplam = 0.0
    for k in kalem_sonuclari:
        ham = kalemler_ham.get(k.kalem_id) or {}
        indirim = None
        try:
            raw = ham.get("indirim_yuzdesi")
            if raw not in (None, ""):
                indirim = max(0.0, min(100.0, float(raw)))
        except (TypeError, ValueError):
            indirim = None

        aylik = float(k.fiyat.aylik_toplam)
        indirimli = round(aylik * (1 - indirim / 100.0), 4) if indirim is not None else None
        katilan = indirimli if indirimli is not None else aylik
        toplam += katilan
        kalem = HesaplamaKalemi(
            hesaplama_id=hesaplama.id,
            urun_tipi=k.urun_tipi,
            ozet=k.urun.ozet(k.yapilandirma, dil),
            aylik_maliyet=aylik,
            indirim_yuzdesi=indirim,
            indirimli_aylik_maliyet=indirimli,
            yapilandirma=k.yapilandirma,
            fiyat_kalemleri=[vars(kalem_kaydi) for kalem_kaydi in k.fiyat.kalemler],
        )
        oturum.add(kalem)

    hesaplama.toplam_aylik_maliyet = toplam
    oturum.add(hesaplama)
    oturum.commit()

    # HTMX: basarili kayittan sonra ilgili listeye yonlendir
    if durum == DURUM_ONAY_BEKLIYOR:
        hedef = "/gecmis/gonderilenler"
        nav_anahtar = "nav_gonderilenler"
        basari_anahtar = "onaya_gonderildi"
    else:
        hedef = "/gecmis/taslaklar"
        nav_anahtar = "nav_tahmin_gecmisi" if ustu_olmayan_mi(kullanici) else "nav_taslaklar"
        basari_anahtar = "tahmin_kaydedildi" if ustu_olmayan_mi(kullanici) else "kaydet_basarili"
    yanit = HTMLResponse(
        f'<div class="ui-alert ui-alert--info">'
        f'"{hesaplama_adi}" {t(basari_anahtar, dil)}'
        f' — <a href="{hedef}">{t(nav_anahtar, dil)}</a>'
        f'</div>'
    )
    yanit.headers["HX-Redirect"] = hedef
    return yanit


@router.post("/disa-aktar")
async def tahmin_disa_aktar(request: Request, kullanici: dict = Depends(yetki_gerekli(IZIN_HESAPLAMA_KULLAN))):
    dil = istekten_dil_al(request)
    form = await request.form()
    genel, kalemler_ham = coklu_kalem_formunu_ayir(form)
    para_birimi = guvenli_para_birimi(genel.get("para_birimi"))

    if not kalemler_ham:
        return HTMLResponse(f'<div class="alert alert-danger">{t("disa_aktar_bos_hata", dil)}</div>', status_code=400)

    kalem_sonuclari = await _tum_kalemleri_coz(kalemler_ham, para_birimi, dil)
    if not kalem_sonuclari or any(k.hata for k in kalem_sonuclari):
        return HTMLResponse(f'<div class="alert alert-danger">{t("fiyat_bulunamadi", dil)}</div>', status_code=400)

    satirlar: list[DisaAktarimSatiri] = []
    genel_toplam = 0.0
    for k in kalem_sonuclari:
        satirlar.extend(k.urun.disa_aktarim_satirlari(k.yapilandirma, k.fiyat, dil))
        genel_toplam += k.fiyat.aylik_toplam

    try:
        icerik = calisma_kitabi_olustur(satirlar, genel_toplam, para_birimi, dil)
    except TahminBosHatasi:
        return HTMLResponse(f'<div class="alert alert-danger">{t("disa_aktar_bos_hata", dil)}</div>', status_code=400)

    dosya_adi = f"azure-tahmin-{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"
    return StreamingResponse(
        io.BytesIO(icerik),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{dosya_adi}"'},
    )


gecmis_router = APIRouter()


def _gecmis_liste_baglami(oturum: Session, kullanici: dict, aktif_sekme: str) -> dict:
    from app.yetkilendirme import ustu_olmayan_mi

    kapsam = gecmis_erisim_kapsami(kullanici)
    tum = oturum.exec(select(Hesaplama).order_by(Hesaplama.olusturulma_tarihi.desc())).all()
    hesaplamalar = [h for h in tum if hesaplamaya_erisebilir_mi(kullanici, h)]

    kullanici_adi_lower = (kullanici.get("kullanici_adi") or "").lower()
    personal_hesaplamalar = [
        h for h in hesaplamalar
        if (h.olusturan_kullanici_adi or "").lower() == kullanici_adi_lower
    ]
    # Ustunde kimse yoksa tum kisisel kayitlar "tahmin gecmisi" listesinde
    if ustu_olmayan_mi(kullanici):
        taslak_hesaplamalar = list(personal_hesaplamalar)
        gonderilen_hesaplamalar = []
    else:
        taslak_hesaplamalar = [
            h for h in personal_hesaplamalar
            if hesaplama_gorunen_durum(h) == "taslak"
        ]
        gonderilen_hesaplamalar = [
            h for h in personal_hesaplamalar
            if hesaplama_gorunen_durum(h) != "taslak"
        ]
    arama_hesaplamalari = [
        h for h in hesaplamalar
        if (h.olusturan_kullanici_adi or "").lower() != kullanici_adi_lower
    ]

    departman_listesi: list[dict] = []
    if kapsam in ("yonetici", "direktor", "admin"):
        dep_anahtarlari: set[str] = set()
        for h in arama_hesaplamalari:
            dep = hesaplama_departmani(h.olusturan_gruplar, h.olusturan_departman)
            if dep:
                dep_anahtarlari.add(dep)
        departman_listesi = sorted(
            [{"anahtar": dep, "etiket": departman_etiketi(dep)} for dep in dep_anahtarlari],
            key=lambda d: d["etiket"],
        )

    return {
        "hesaplamalar": hesaplamalar,
        "personal_hesaplamalar": personal_hesaplamalar,
        "taslak_hesaplamalar": taslak_hesaplamalar,
        "gonderilen_hesaplamalar": gonderilen_hesaplamalar,
        "arama_hesaplamalari": arama_hesaplamalari,
        "departman_listesi": departman_listesi,
        "gecmis_gorunumu": kapsam,
        "aktif_sekme": aktif_sekme,
        "ustu_yok": ustu_olmayan_mi(kullanici),
    }


@gecmis_router.get("/gecmis")
async def gecmis_listesi(
    request: Request,
    kullanici: dict = Depends(gecmis_goruntule_gerekli),
):
    from app.yetkilendirme import IZIN_HESAPLAMA_KULLAN, kullanici_izinli_mi

    if not kullanici_izinli_mi(kullanici, IZIN_HESAPLAMA_KULLAN):
        return RedirectResponse("/gecmis/arama", status_code=303)
    return RedirectResponse("/gecmis/taslaklar", status_code=303)


@gecmis_router.get("/gecmis/taslaklar")
async def gecmis_taslaklar(
    request: Request,
    oturum: Session = Depends(oturum_al),
    kullanici: dict = Depends(gecmis_goruntule_gerekli),
):
    return render(request, "gecmis.html", _gecmis_liste_baglami(oturum, kullanici, "taslaklar"))


@gecmis_router.get("/gecmis/gonderilenler")
async def gecmis_gonderilenler(
    request: Request,
    oturum: Session = Depends(oturum_al),
    kullanici: dict = Depends(gecmis_goruntule_gerekli),
):
    from app.yetkilendirme import ustu_olmayan_mi

    if ustu_olmayan_mi(kullanici):
        return RedirectResponse("/gecmis/taslaklar", status_code=303)
    return render(request, "gecmis.html", _gecmis_liste_baglami(oturum, kullanici, "gonderilenler"))


@gecmis_router.get("/gecmis/arama")
async def gecmis_arama(
    request: Request,
    oturum: Session = Depends(oturum_al),
    kullanici: dict = Depends(gecmis_goruntule_gerekli),
):
    kapsam = gecmis_erisim_kapsami(kullanici)
    if kapsam not in ("yonetici", "direktor", "admin"):
        return RedirectResponse("/gecmis/taslaklar", status_code=303)
    return render(request, "gecmis.html", _gecmis_liste_baglami(oturum, kullanici, "arama"))


@gecmis_router.get("/gecmis/{hesaplama_id}")
async def gecmis_detay(
    hesaplama_id: int,
    request: Request,
    oturum: Session = Depends(oturum_al),
    kullanici: dict = Depends(gecmis_goruntule_gerekli),
):
    dil = istekten_dil_al(request)
    hesaplama = oturum.get(Hesaplama, hesaplama_id)
    if hesaplama is None:
        return HTMLResponse(f'<div class="alert alert-danger">{t("gecmis_bulunamadi", dil)}</div>', status_code=404)
    if not hesaplamaya_erisebilir_mi(kullanici, hesaplama):
        return HTMLResponse(
            f'<div class="alert alert-danger">{t("gecmis_yalnizca_sahibi_erisebilir", dil)}</div>', status_code=403
        )

    from app.yetkilendirme import hesaplamayi_iptal_edebilir_mi

    return render(
        request,
        "gecmis_detay.html",
        {
            "hesaplama": hesaplama,
            "gecmis_gorunumu": gecmis_erisim_kapsami(kullanici),
            "iptal_edebilir": hesaplamayi_iptal_edebilir_mi(kullanici, hesaplama),
            "duzenleyebilir": hesaplamayi_duzenleyebilir_mi(kullanici, hesaplama),
            "gorunen_durum": hesaplama_gorunen_durum(hesaplama),
        },
    )


def _hesaplamadan_satirlar(hesaplama, dil: str = "en") -> tuple[list[DisaAktarimSatiri], float]:
    """Kaydedilmis Hesaplama'dan Azure formatinda DisaAktarimSatiri listesi uretir."""
    return hesaplamadan_satirlar(hesaplama, dil)


@gecmis_router.get("/gecmis/{hesaplama_id}/excel")
async def gecmis_detay_excel(
    hesaplama_id: int,
    request: Request,
    oturum: Session = Depends(oturum_al),
    kullanici: dict = Depends(gecmis_goruntule_gerekli),
):
    dil = istekten_dil_al(request)
    hesaplama = oturum.get(Hesaplama, hesaplama_id)
    if hesaplama is None or not hesaplamaya_erisebilir_mi(kullanici, hesaplama):
        return HTMLResponse(f'<div class="alert alert-danger">{t("gecmis_bulunamadi", dil)}</div>', status_code=404)
    satirlar, toplam = _hesaplamadan_satirlar(hesaplama, dil)
    from app.disa_aktar import hesaplama_meta_olustur

    try:
        icerik = calisma_kitabi_olustur(
            satirlar,
            toplam,
            hesaplama.para_birimi,
            dil,
            meta=hesaplama_meta_olustur(hesaplama),
        )
    except TahminBosHatasi:
        return HTMLResponse(f'<div class="alert alert-danger">{t("disa_aktar_bos_hata", dil)}</div>', status_code=400)
    dosya_adi = f"azure-tahmin-{hesaplama.ad}-{hesaplama.olusturulma_tarihi.strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        io.BytesIO(icerik),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{dosya_adi}"'},
    )


@gecmis_router.get("/gecmis-excel")
async def gecmis_tumu_excel(
    request: Request,
    oturum: Session = Depends(oturum_al),
    kullanici: dict = Depends(gecmis_goruntule_gerekli),
):
    """Kullanicinin gorebilecegi tum hesaplamalari tek bir Excel dosyasina aktarir."""
    dil = istekten_dil_al(request)
    from app.disa_aktar import hesaplama_meta_olustur
    from openpyxl import load_workbook

    tum = oturum.exec(select(Hesaplama).order_by(Hesaplama.olusturulma_tarihi.desc())).all()
    hesaplamalar = [h for h in tum if hesaplamaya_erisebilir_mi(kullanici, h)]

    kitap = None
    for hesaplama in hesaplamalar:
        satirlar, toplam = _hesaplamadan_satirlar(hesaplama, dil)
        if not satirlar:
            continue
        icerik = calisma_kitabi_olustur(
            satirlar,
            toplam,
            hesaplama.para_birimi,
            dil,
            meta=hesaplama_meta_olustur(hesaplama),
        )
        tek_kitap = load_workbook(io.BytesIO(icerik))
        if kitap is None:
            kitap = tek_kitap
            sayfa_adi = (hesaplama.ad or str(hesaplama.id))[:31]
            kitap.active.title = sayfa_adi
        else:
            sayfa_adi = (hesaplama.ad or str(hesaplama.id))[:31]
            mevcut = set(kitap.sheetnames)
            temel = sayfa_adi
            i = 2
            while sayfa_adi in mevcut:
                sayfa_adi = f"{temel[:28]}_{i}"
                i += 1
            kaynak = tek_kitap.active
            yeni = kitap.create_sheet(title=sayfa_adi)
            for row in kaynak.iter_rows():
                for cell in row:
                    yeni[cell.coordinate].value = cell.value
                    if cell.has_style:
                        yeni[cell.coordinate]._style = cell._style
            for col, dim in kaynak.column_dimensions.items():
                yeni.column_dimensions[col].width = dim.width

    if kitap is None:
        from openpyxl import Workbook
        kitap = Workbook()
        kitap.active.title = "Bos"

    arabellek = io.BytesIO()
    kitap.save(arabellek)
    dosya_adi = f"azure-tahminler-{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"
    return StreamingResponse(
        io.BytesIO(arabellek.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{dosya_adi}"'},
    )


@gecmis_router.post("/gecmis/{hesaplama_id}/sil")
async def gecmis_sil(
    hesaplama_id: int,
    request: Request,
    oturum: Session = Depends(oturum_al),
    kullanici: dict = Depends(gecmis_goruntule_gerekli),
):
    dil = istekten_dil_al(request)
    hesaplama = oturum.get(Hesaplama, hesaplama_id)
    if hesaplama is None:
        return HTMLResponse(f'<div class="alert alert-danger">{t("gecmis_bulunamadi", dil)}</div>', status_code=404)
    if not hesaplamayi_silebilir_mi(kullanici, hesaplama):
        return HTMLResponse(
            f'<div class="alert alert-danger">{t("gecmis_yalnizca_sahibi_erisebilir", dil)}</div>', status_code=403
        )

    oturum.delete(hesaplama)
    oturum.commit()
    return RedirectResponse("/gecmis/taslaklar", status_code=303)
