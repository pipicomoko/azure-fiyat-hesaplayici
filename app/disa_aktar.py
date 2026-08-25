"""Tahminin .xlsx (Excel) olarak disa aktarilmasi.

Export Tur A: Azure orijinal format + AFH ust bilgi / indirim / yillik sutunlari.
Export Tur B: Donemsel ozet (onaylanmis kayitlar).
"""

from __future__ import annotations

import io
import re
import textwrap
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.i18n import Dil
from app.products.base import DisaAktarimSatiri


class TahminBosHatasi(Exception):
    """Bos bir tahmin disa aktarilamaz."""


# openpyxl worksheet title: max 31 char; yasak: \ / * ? : [ ]
_EXCEL_SAYFA_YASAK = re.compile(r"[\\/*?:\[\]]")


def excel_sayfa_adi(ad: str | None, yedek: str = "Sayfa") -> str:
    """openpyxl'in kabul ettigi guvenli sayfa adi (BUG-02)."""
    ham = (ad or "").strip() or str(yedek)
    temiz = _EXCEL_SAYFA_YASAK.sub("_", ham).strip("'").strip()
    if not temiz:
        temiz = str(yedek) or "Sayfa"
    return temiz[:31]


def benzersiz_excel_sayfa_adi(
    ad: str | None, mevcut: set[str], yedek: str = "Sayfa"
) -> str:
    """Yasak karakterleri temizler ve workbook icinde tekil tutar."""
    temel = excel_sayfa_adi(ad, yedek)
    aday = temel
    i = 2
    while aday in mevcut:
        ek = f"_{i}"
        aday = f"{temel[: max(1, 31 - len(ek))]}{ek}"
        i += 1
    return aday


_BASLIK_DOLGU = PatternFill("solid", fgColor="F2F2F2")
_ACIKLAMA_GENISLIGI = 110
_SATIR_YUKSEKLIGI = 15.0
_DISCLAIMER = (
    "All prices shown are in United States – Dollar ($). "
    "This is an estimate and is not a quote. Prices are subject to change."
)
# Literal "$" so TR Excel does not substitute ₺; ","/"." follow workbook locale.
_MALIYET_BICIMI = '"$" #,##0.00'
_BOS_MALIYET = "—"

_COL_AYLIK = 6
_COL_INDIRIM = 7
_COL_INDIRIMLI_AYLIK = 8
_COL_INDIRIMLI_YILLIK = 9
_COL_ON_ODEME = 10
_COL_YILLIK = 11


@dataclass
class HesaplamaMeta:
    senaryo_adi: str = ""
    olusturan: str = ""
    birim: str = ""
    olusturma_tarihi: str = ""
    durum: str = ""
    revizyon: int = 1
    onaylayan: str = ""
    onay_tarihi: str = ""
    reddeden: str = ""


def hesaplama_meta_olustur(hesaplama) -> HesaplamaMeta:
    from app.yetkilendirme import (
        departman_etiketi,
        hesaplama_departmani,
        hesaplama_reddeden_sam,
        sam_gorunen_adi,
    )

    dep = hesaplama_departmani(
        getattr(hesaplama, "olusturan_gruplar", None),
        getattr(hesaplama, "olusturan_departman", None),
    )
    onay_t = getattr(hesaplama, "onay_tarihi", None)
    olusturma = getattr(hesaplama, "olusturulma_tarihi", None)
    reddeden_sam = hesaplama_reddeden_sam(hesaplama)
    return HesaplamaMeta(
        senaryo_adi=str(getattr(hesaplama, "ad", "") or ""),
        olusturan=str(
            getattr(hesaplama, "olusturan_ad_soyad", None)
            or getattr(hesaplama, "olusturan_kullanici_adi", "")
            or ""
        ),
        birim=departman_etiketi(dep) if dep else "",
        olusturma_tarihi=olusturma.strftime("%Y-%m-%d %H:%M") if olusturma else "",
        durum=str(getattr(hesaplama, "durum", "") or ""),
        revizyon=int(getattr(hesaplama, "revizyon", 1) or 1),
        onaylayan=str(getattr(hesaplama, "onaylayan_kullanici_adi", "") or ""),
        onay_tarihi=onay_t.strftime("%Y-%m-%d %H:%M") if onay_t else "",
        reddeden=sam_gorunen_adi(reddeden_sam) if reddeden_sam else "",
    )


def _satir_yuksekligi(metin: str | None, sutun_genisligi: int) -> float:
    if not metin:
        return _SATIR_YUKSEKLIGI
    satir_sayisi = 0
    for paragraf in str(metin).split("\n"):
        sarilmis = textwrap.wrap(paragraf, width=sutun_genisligi) or [""]
        satir_sayisi += len(sarilmis)
    return max(_SATIR_YUKSEKLIGI, satir_sayisi * _SATIR_YUKSEKLIGI)


def _hucrele(
    ws,
    row: int,
    col: int,
    value,
    bold=False,
    fill=None,
    align="left",
    number_format=None,
):
    cell = ws.cell(row=row, column=col, value=value)
    if bold:
        cell.font = Font(bold=True)
    if fill:
        cell.fill = fill
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    if number_format:
        cell.number_format = number_format
    return cell


def _maliyet_hucresi(ws, row: int, col: int, value, bold=False):
    return _hucrele(
        ws,
        row,
        col,
        round(float(value), 2),
        bold=bold,
        align="right",
        number_format=_MALIYET_BICIMI,
    )


def indirim_yuzdesini_oku(ham) -> float | None:
    """Form/kayit ham degerinden 0–100 indirim yuzdesi; gecersizse None."""
    if ham in (None, ""):
        return None
    try:
        deger = float(ham)
    except (TypeError, ValueError):
        return None
    if deger < 0 or deger > 100:
        return None
    return deger


def indirimli_aylik_hesapla(
    aylik: float, indirim_yuzdesi: float | None
) -> float | None:
    """Web kayit ile ayni: aylik * (1 - indirim/100), 4 hane."""
    if indirim_yuzdesi is None:
        return None
    return round(float(aylik) * (1 - float(indirim_yuzdesi) / 100.0), 4)


def satirlara_indirim_uygula(
    satirlar: list[DisaAktarimSatiri],
    aylik: float,
    indirim_ham,
) -> float:
    """Indirimi export satirlarina yazar; toplama katilan aylik tutari dondurur."""
    indirim = indirim_yuzdesini_oku(indirim_ham)
    indirimli = indirimli_aylik_hesapla(aylik, indirim)
    if indirim is None:
        return float(aylik)
    for satir in satirlar:
        satir.indirim_yuzdesi = indirim
        satir.indirimli_aylik = indirimli
    return float(indirimli) if indirimli is not None else float(aylik)


def _tahmin_sayfasini_doldur(
    ws,
    satirlar: list[DisaAktarimSatiri],
    genel_toplam: float,
    para_birimi: str,
    dil: Dil,
    meta: HesaplamaMeta | None = None,
) -> None:
    """Tek bir worksheet'e tahmin satirlari yazar (coklu export icin yeniden kullanilir)."""
    _hucrele(ws, 1, 1, "Microsoft Azure Estimate", bold=True)
    ws.row_dimensions[1].height = 18

    row = 2
    if meta:
        _hucrele(ws, row, 1, meta.senaryo_adi or "Your Estimate")
        row += 1
        for etiket, deger in (
            ("Olusturan Calisan", meta.olusturan),
            ("Departman", meta.birim),
            ("Olusturulma Tarihi", meta.olusturma_tarihi),
            ("Durum", f"{meta.durum} / Rev {meta.revizyon}"),
            ("Reddeden", meta.reddeden),
            ("Onaylayan", meta.onaylayan),
            ("Onay Tarihi", meta.onay_tarihi),
        ):
            if not deger:
                continue
            _hucrele(ws, row, 1, etiket, bold=True)
            _hucrele(ws, row, 2, deger)
            row += 1
    else:
        _hucrele(ws, row, 1, "Your Estimate")
        row += 1

    baslik_satiri = row
    basliklar = [
        "Service category",
        "Service type",
        "Custom name",
        "Region",
        "Description",
        "Estimated monthly cost",
        "Indirim Yuzdesi",
        "Indirimli Aylik Maliyet",
        "İndirimli Yıllık Maliyeti",
        "Estimated upfront cost",
        "Yillik Tahmini Maliyet",
    ]
    for col, baslik in enumerate(basliklar, 1):
        _hucrele(
            ws,
            baslik_satiri,
            col,
            baslik,
            bold=True,
            fill=_BASLIK_DOLGU,
            align="center" if col >= 6 else "left",
        )
    ws.row_dimensions[baslik_satiri].height = 15

    indirimli_yillik_toplam = 0.0
    liste_aylik_toplam = 0.0
    for satir in satirlar:
        r = ws.max_row + 1
        aylik = float(satir.ara_toplam)
        liste_aylik_toplam += aylik
        indirim = getattr(satir, "indirim_yuzdesi", None)
        indirimli = getattr(satir, "indirimli_aylik", None)
        _hucrele(ws, r, 1, satir.servis_kategori or "")
        _hucrele(ws, r, 2, satir.urun)
        _hucrele(ws, r, 3, satir.ozel_ad or "")
        _hucrele(ws, r, 4, satir.bolge)
        _hucrele(ws, r, 5, satir.yapilandirma_ozeti)
        _maliyet_hucresi(ws, r, _COL_AYLIK, aylik)
        if indirim is not None:
            _hucrele(
                ws,
                r,
                _COL_INDIRIM,
                round(float(indirim), 2),
                align="right",
                number_format="0.00",
            )
            ind_aylik = float(indirimli) if indirimli is not None else aylik
            # Yalniz bu sutun indirimli yillik: indirimli_aylik * 12
            ind_yillik = round(ind_aylik * 12, 2)
            indirimli_yillik_toplam += ind_yillik
            _maliyet_hucresi(ws, r, _COL_INDIRIMLI_AYLIK, ind_aylik)
            _maliyet_hucresi(ws, r, _COL_INDIRIMLI_YILLIK, ind_yillik)
        else:
            _hucrele(ws, r, _COL_INDIRIM, _BOS_MALIYET)
            _hucrele(ws, r, _COL_INDIRIMLI_AYLIK, _BOS_MALIYET)
            _hucrele(ws, r, _COL_INDIRIMLI_YILLIK, _BOS_MALIYET)
        _maliyet_hucresi(ws, r, _COL_ON_ODEME, satir.on_odeme)
        # Yillik Tahmini Maliyet her zaman liste aylik * 12 (indirimsiz)
        _maliyet_hucresi(ws, r, _COL_YILLIK, aylik * 12)
        ws.row_dimensions[r].height = _satir_yuksekligi(
            satir.yapilandirma_ozeti, _ACIKLAMA_GENISLIGI
        )

    r = ws.max_row + 1
    _hucrele(ws, r, 1, "Support")
    _hucrele(ws, r, 2, "Support")
    _hucrele(ws, r, 3, "")
    _hucrele(ws, r, 4, "Support")
    _hucrele(ws, r, 5, "")
    _maliyet_hucresi(ws, r, _COL_AYLIK, 0)
    _hucrele(ws, r, _COL_INDIRIM, "")
    _hucrele(ws, r, _COL_INDIRIMLI_AYLIK, "")
    _hucrele(ws, r, _COL_INDIRIMLI_YILLIK, "")
    _maliyet_hucresi(ws, r, _COL_ON_ODEME, 0)
    _maliyet_hucresi(ws, r, _COL_YILLIK, 0)

    r = ws.max_row + 1
    _hucrele(ws, r, 4, "Licensing Program")
    _hucrele(ws, r, 5, "Microsoft Customer Agreement (MCA)")
    r = ws.max_row + 1
    _hucrele(ws, r, 4, "Billing Account")
    _hucrele(ws, r, 5, "")
    r = ws.max_row + 1
    _hucrele(ws, r, 4, "Billing Profile")
    _hucrele(ws, r, 5, "")

    r = ws.max_row + 1
    _hucrele(ws, r, 4, "Total", bold=True)
    _maliyet_hucresi(ws, r, _COL_AYLIK, genel_toplam, bold=True)
    _maliyet_hucresi(ws, r, _COL_INDIRIMLI_YILLIK, indirimli_yillik_toplam, bold=True)
    _maliyet_hucresi(ws, r, _COL_ON_ODEME, 0, bold=True)
    _maliyet_hucresi(ws, r, _COL_YILLIK, liste_aylik_toplam * 12, bold=True)

    r = ws.max_row + 2
    _hucrele(ws, r, 1, _DISCLAIMER)
    r = ws.max_row + 1
    _hucrele(
        ws,
        r,
        1,
        f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
    )

    sutun_genislikleri = {
        1: 18,
        2: 20,
        3: 16,
        4: 18,
        5: _ACIKLAMA_GENISLIGI,
        _COL_AYLIK: 22,
        _COL_INDIRIM: 14,
        _COL_INDIRIMLI_AYLIK: 22,
        _COL_INDIRIMLI_YILLIK: 26,
        _COL_ON_ODEME: 22,
        _COL_YILLIK: 22,
    }
    for col, genislik in sutun_genislikleri.items():
        ws.column_dimensions[get_column_letter(col)].width = genislik

    for row_obj in ws.iter_rows(min_row=baslik_satiri + 1, max_col=5, min_col=5):
        for cell in row_obj:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def calisma_kitabi_olustur(
    satirlar: list[DisaAktarimSatiri],
    genel_toplam: float,
    para_birimi: str,
    dil: Dil,
    meta: HesaplamaMeta | None = None,
) -> bytes:
    if not satirlar:
        raise TahminBosHatasi()

    kitap = Workbook()
    ws = kitap.active
    ws.title = "Your Estimate"
    _tahmin_sayfasini_doldur(ws, satirlar, genel_toplam, para_birimi, dil, meta)
    arabellek = io.BytesIO()
    kitap.save(arabellek)
    return arabellek.getvalue()


@dataclass(frozen=True)
class CokluExcelPaket:
    """Toplu Excel icin onceden cozumlenmis (ORM-bagimsiz) paket."""

    ad: str
    kayit_id: int | str
    satirlar: list[DisaAktarimSatiri]
    genel_toplam: float
    para_birimi: str
    dil: Dil
    meta: HesaplamaMeta | None = None


def coklu_calisma_kitabi_olustur(paketler: list[CokluExcelPaket]) -> bytes:
    """Birden fazla tahmini TEK workbook'ta, sayfa basina yazarak uretir (BUG-06).

    Eski yol her kayit icin workbook serialize+load+hucre kopyalama yapiyordu;
    bu yol dogrudan sayfa doldurur (CPU yogun is async thread'e tasinmali).
    """
    kitap = Workbook()
    mevcut: set[str] = set()
    ilk = True
    for paket in paketler:
        if not paket.satirlar:
            continue
        sayfa_adi = benzersiz_excel_sayfa_adi(
            paket.ad, mevcut, yedek=str(paket.kayit_id)
        )
        mevcut.add(sayfa_adi)
        if ilk:
            ws = kitap.active
            ws.title = sayfa_adi
            ilk = False
        else:
            ws = kitap.create_sheet(title=sayfa_adi)
        _tahmin_sayfasini_doldur(
            ws,
            paket.satirlar,
            paket.genel_toplam,
            paket.para_birimi,
            paket.dil,
            paket.meta,
        )

    if ilk:
        kitap.active.title = "Bos"

    arabellek = io.BytesIO()
    kitap.save(arabellek)
    return arabellek.getvalue()


def donemsel_rapor_kitabi_olustur(hesaplamalar: list[Any]) -> bytes:
    """Export Tur B — her satir bir onaylanmis hesaplama ozeti."""
    from app.yetkilendirme import departman_etiketi, hesaplama_departmani

    kitap = Workbook()
    ws = kitap.active
    ws.title = "Donemsel Rapor"
    basliklar = [
        "Tarih",
        "Calisan",
        "Departman",
        "Hizmet",
        "Kalem Sayisi",
        "Toplam Tutar",
        "Onaylayan",
        "Onay Tarihi",
    ]
    for col, baslik in enumerate(basliklar, 1):
        _hucrele(ws, 1, col, baslik, bold=True, fill=_BASLIK_DOLGU)

    for h in hesaplamalar:
        dep = hesaplama_departmani(h.olusturan_gruplar, h.olusturan_departman)
        hizmetler = sorted({(k.urun_tipi or "") for k in (h.kalemler or [])})
        onay_t = h.onay_tarihi
        olusturma = h.olusturulma_tarihi
        r = ws.max_row + 1
        _hucrele(ws, r, 1, olusturma.strftime("%Y-%m-%d") if olusturma else "")
        _hucrele(ws, r, 2, h.olusturan_ad_soyad or h.olusturan_kullanici_adi or "")
        _hucrele(ws, r, 3, departman_etiketi(dep) if dep else "")
        _hucrele(ws, r, 4, ", ".join(hizmetler) or h.ad)
        _hucrele(ws, r, 5, len(h.kalemler or []))
        _hucrele(
            ws,
            r,
            6,
            round(float(h.toplam_aylik_maliyet or 0), 2),
            align="right",
            number_format="#,##0.00",
        )
        _hucrele(ws, r, 7, h.onaylayan_kullanici_adi or "")
        _hucrele(ws, r, 8, onay_t.strftime("%Y-%m-%d") if onay_t else "")

    for col, genislik in enumerate([12, 22, 16, 28, 12, 14, 18, 14], 1):
        ws.column_dimensions[get_column_letter(col)].width = genislik

    arabellek = io.BytesIO()
    kitap.save(arabellek)
    return arabellek.getvalue()
