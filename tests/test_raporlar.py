"""Raporlar secimli Excel export (UI secim + ids filtresi)."""

from __future__ import annotations

import re
from datetime import datetime, timezone
from urllib.parse import unquote

from sqlmodel import Session

from app.main import app
from app.models import DURUM_ONAYLANDI, Hesaplama, HesaplamaKalemi
from app.yetkilendirme import aktif_kullanici

_YONETICI = {
    "kullanici_adi": "onur.simsek",
    "ad_soyad": "Onur Simsek",
    "unvan": "Mudur",
    "gruplar": ["AFH-Calisanlar", "AFH-Yoneticiler"],
    "manager": "emre.turan",
    "manager_zinciri": ["emre.turan"],
    "rol": "yonetici",
}


def _onayli(oturum: Session, ad: str) -> Hesaplama:
    h = Hesaplama(
        ad=ad,
        durum=DURUM_ONAYLANDI,
        toplam_aylik_maliyet=12.5,
        para_birimi="USD",
        olusturan_kullanici_adi="test.kullanici",
        olusturan_ad_soyad="Test Kullanici",
        olusturan_gruplar=["AFH-Calisanlar"],
        olusturan_departman="finans",
        olusturan_manager_zinciri=["onur.simsek", "emre.turan"],
        onaylayan_kullanici_adi="onur.simsek",
        onay_tarihi=datetime.now(timezone.utc),
    )
    oturum.add(h)
    oturum.commit()
    oturum.refresh(h)
    oturum.add(
        HesaplamaKalemi(
            hesaplama_id=h.id,
            urun_tipi="managed_disks",
            ozet="disk",
            aylik_maliyet=12.5,
            yapilandirma={"bolge": "eastus"},
            fiyat_kalemleri=[],
        )
    )
    oturum.commit()
    return h


def test_raporlar_sayfasi_export_cubugu_ve_checkbox(client, veritabani):
    app.dependency_overrides[aktif_kullanici] = lambda: _YONETICI
    with Session(veritabani) as oturum:
        _onayli(oturum, "Rapor A")

    yanit = client.get("/raporlar")
    assert yanit.status_code == 200
    assert "data-rapor-export" in yanit.text
    assert "data-rapor-hepsini" in yanit.text
    assert "data-rapor-export-secilen" in yanit.text
    assert "data-rapor-sec" in yanit.text
    assert "disabled" in yanit.text


def test_raporlar_excel_secili_idler(client, veritabani):
    app.dependency_overrides[aktif_kullanici] = lambda: _YONETICI
    with Session(veritabani) as oturum:
        a = _onayli(oturum, "Secili A")
        b = _onayli(oturum, "Secili B")
        a_id, b_id = a.id, b.id

    yanit = client.get(f"/raporlar/excel?ids={a_id}")
    assert yanit.status_code == 200
    assert len(yanit.content) > 100
    cd = yanit.headers.get("content-disposition", "")
    assert "afh-rapor-" not in cd
    utf8_adi = unquote(cd.split("filename*=UTF-8''", 1)[1])
    assert re.fullmatch(r"azure-tahminler-\d{8}-\d{4}\.xlsx", utf8_adi)

    from io import BytesIO

    from openpyxl import load_workbook

    kitap = load_workbook(BytesIO(yanit.content))
    ws = kitap.active
    assert ws.max_row == 2

    yanit2 = client.get(f"/raporlar/excel?ids={a_id}&ids={b_id}")
    assert yanit2.status_code == 200
    kitap2 = load_workbook(BytesIO(yanit2.content))
    assert kitap2.active.max_row == 3


def test_gecmis_arama_export_cubugu(client, veritabani):
    app.dependency_overrides[aktif_kullanici] = lambda: _YONETICI
    with Session(veritabani) as oturum:
        h = _onayli(oturum, "Arama Kaydi")
        h.olusturan_manager_zinciri = ["onur.simsek", "emre.turan"]
        oturum.add(h)
        oturum.commit()

    yanit = client.get("/gecmis/arama")
    assert yanit.status_code == 200
    assert 'data-export-url="/gecmis-excel"' in yanit.text
    assert "data-rapor-sec" in yanit.text


def test_gecmis_excel_secili_idler(client, veritabani):
    app.dependency_overrides[aktif_kullanici] = lambda: _YONETICI
    with Session(veritabani) as oturum:
        a = _onayli(oturum, "Gecmis A")
        a.olusturan_manager_zinciri = ["onur.simsek"]
        oturum.add(a)
        oturum.commit()
        oturum.refresh(a)
        a_id = a.id

    yanit = client.get(f"/gecmis-excel?ids={a_id}")
    assert yanit.status_code == 200
    assert len(yanit.content) > 100
    cd = yanit.headers.get("content-disposition", "")
    utf8_adi = unquote(cd.split("filename*=UTF-8''", 1)[1])
    assert re.fullmatch(r"azure-tahminler-\d{8}-\d{4}\.xlsx", utf8_adi)


def test_raporlar_sayfalama_sinirlari(client, veritabani):
    app.dependency_overrides[aktif_kullanici] = lambda: _YONETICI
    with Session(veritabani) as oturum:
        for i in range(21):
            _onayli(oturum, f"Sayfa {i}")

    birinci = client.get("/raporlar?sayfa=1")
    assert birinci.status_code == 200
    assert "1 / 2" in birinci.text
    ikinci = client.get("/raporlar?sayfa=2")
    assert "Sayfa 0" in ikinci.text
    tasma = client.get("/raporlar?sayfa=99")
    assert tasma.status_code == 200
    assert "2 / 2" in tasma.text
    gecersiz = client.get("/raporlar?sayfa=abc")
    assert gecersiz.status_code == 200


def test_gecmis_arama_filtre_kisi(client, veritabani):
    app.dependency_overrides[aktif_kullanici] = lambda: _YONETICI
    with Session(veritabani) as oturum:
        a = _onayli(oturum, "Ahmet Kaydi")
        a.olusturan_kullanici_adi = "ahmet.yildirim"
        a.olusturan_ad_soyad = "Ahmet Yildirim"
        a.olusturan_manager_zinciri = ["onur.simsek"]
        b = _onayli(oturum, "Elif Kaydi")
        b.olusturan_kullanici_adi = "elif.aydin"
        b.olusturan_ad_soyad = "Elif Aydin"
        b.olusturan_manager_zinciri = ["onur.simsek"]
        oturum.add(a)
        oturum.add(b)
        oturum.commit()

    yanit = client.get("/gecmis/arama?kisi=elif")
    assert yanit.status_code == 200
    assert "Elif Kaydi" in yanit.text
    assert "Ahmet Kaydi" not in yanit.text


def test_raporlar_varsayilan_tarih_araligi_eskiyi_gizler(client, veritabani):
    app.dependency_overrides[aktif_kullanici] = lambda: _YONETICI
    from app.tarih_filtre import varsayilan_tarih_iso

    with Session(veritabani) as oturum:
        _onayli(oturum, "Yeni Rapor")
        eski = _onayli(oturum, "Eski Rapor")
        eski.onay_tarihi = datetime(2025, 1, 10, tzinfo=timezone.utc)
        oturum.add(eski)
        oturum.commit()

    bas, bit = varsayilan_tarih_iso()
    yanit = client.get("/raporlar")
    assert yanit.status_code == 200
    assert f'name="baslangic" value="{bas}"' in yanit.text
    assert f'name="bitis" value="{bit}"' in yanit.text
    assert "Yeni Rapor" in yanit.text
    assert "Eski Rapor" not in yanit.text


def test_raporlar_acik_tarih_parametreleri_korunur(client, veritabani):
    app.dependency_overrides[aktif_kullanici] = lambda: _YONETICI
    with Session(veritabani) as oturum:
        _onayli(oturum, "Yeni Rapor")
        eski = _onayli(oturum, "Eski Rapor")
        eski.onay_tarihi = datetime(2025, 1, 10, tzinfo=timezone.utc)
        oturum.add(eski)
        oturum.commit()

    yanit = client.get("/raporlar?baslangic=2025-01-01&bitis=2025-01-31")
    assert yanit.status_code == 200
    assert 'name="baslangic" value="2025-01-01"' in yanit.text
    assert 'name="bitis" value="2025-01-31"' in yanit.text
    assert "Eski Rapor" in yanit.text
    assert "Yeni Rapor" not in yanit.text


def test_gecmis_arama_varsayilan_ve_acik_tarih(client, veritabani):
    app.dependency_overrides[aktif_kullanici] = lambda: _YONETICI
    from app.tarih_filtre import varsayilan_tarih_iso

    with Session(veritabani) as oturum:
        yeni = _onayli(oturum, "Yeni Arama")
        yeni.olusturan_kullanici_adi = "elif.aydin"
        yeni.olusturan_manager_zinciri = ["onur.simsek"]
        eski = _onayli(oturum, "Eski Arama")
        eski.olusturan_kullanici_adi = "elif.aydin"
        eski.olusturan_manager_zinciri = ["onur.simsek"]
        eski.olusturulma_tarihi = datetime(2025, 1, 10, tzinfo=timezone.utc)
        oturum.add(yeni)
        oturum.add(eski)
        oturum.commit()

    bas, bit = varsayilan_tarih_iso()
    ilk = client.get("/gecmis/arama")
    assert ilk.status_code == 200
    assert f'name="baslangic" value="{bas}"' in ilk.text
    assert f'name="bitis" value="{bit}"' in ilk.text
    assert "Yeni Arama" in ilk.text
    assert "Eski Arama" not in ilk.text

    acik = client.get("/gecmis/arama?baslangic=2025-01-01&bitis=2025-01-31")
    assert "Eski Arama" in acik.text
    assert "Yeni Arama" not in acik.text
    assert 'name="baslangic" value="2025-01-01"' in acik.text


def test_onay_kuyrugu_varsayilan_ve_acik_tarih(client, veritabani):
    app.dependency_overrides[aktif_kullanici] = lambda: _YONETICI
    from app.models import DURUM_ONAY_BEKLIYOR
    from app.tarih_filtre import varsayilan_tarih_iso

    def _bekleyen(oturum: Session, ad: str) -> Hesaplama:
        h = Hesaplama(
            ad=ad,
            durum=DURUM_ONAY_BEKLIYOR,
            toplam_aylik_maliyet=10,
            para_birimi="USD",
            olusturan_kullanici_adi="elif.aydin",
            olusturan_ad_soyad="Elif Aydin",
            olusturan_departman="finans",
            olusturan_manager_zinciri=["onur.simsek"],
            onay_hedefi="onur.simsek",
        )
        oturum.add(h)
        oturum.commit()
        oturum.refresh(h)
        return h

    with Session(veritabani) as oturum:
        _bekleyen(oturum, "Yeni Onay")
        eski = _bekleyen(oturum, "Eski Onay")
        eski.olusturulma_tarihi = datetime(2025, 1, 10, tzinfo=timezone.utc)
        oturum.add(eski)
        oturum.commit()

    bas, bit = varsayilan_tarih_iso()
    ilk = client.get("/onay-kuyrugu")
    assert ilk.status_code == 200
    assert f'name="baslangic" value="{bas}"' in ilk.text
    assert f'name="bitis" value="{bit}"' in ilk.text
    assert "Yeni Onay" in ilk.text
    assert "Eski Onay" not in ilk.text

    acik = client.get("/onay-kuyrugu?baslangic=2025-01-01&bitis=2025-01-31")
    assert "Eski Onay" in acik.text
    assert "Yeni Onay" not in acik.text


def test_gecmis_taslaklar_varsayilan_tarih_inputlari(client, veritabani):
    from app.models import DURUM_TASLAK
    from app.tarih_filtre import varsayilan_tarih_iso

    with Session(veritabani) as oturum:
        oturum.add(
            Hesaplama(
                ad="Kisisel Taslak",
                durum=DURUM_TASLAK,
                toplam_aylik_maliyet=5,
                para_birimi="USD",
                olusturan_kullanici_adi="test.kullanici",
            )
        )
        oturum.commit()

    bas, bit = varsayilan_tarih_iso()
    yanit = client.get("/gecmis/taslaklar")
    assert yanit.status_code == 200
    assert f'id="personalTarihBaslangic" value="{bas}"' in yanit.text
    assert f'id="personalTarihBitis" value="{bit}"' in yanit.text
    assert "Kisisel Taslak" in yanit.text
    assert "personalFiltreKur" in yanit.text
    assert "filtrele();" in yanit.text


def test_gecmis_gonderilenler_varsayilan_tarih_inputlari(client, veritabani):
    from app.models import DURUM_ONAY_BEKLIYOR
    from app.tarih_filtre import varsayilan_tarih_iso

    with Session(veritabani) as oturum:
        oturum.add(
            Hesaplama(
                ad="Kisisel Gonderilen",
                durum=DURUM_ONAY_BEKLIYOR,
                toplam_aylik_maliyet=8,
                para_birimi="USD",
                olusturan_kullanici_adi="test.kullanici",
                onay_hedefi="onur.simsek",
            )
        )
        oturum.commit()

    bas, bit = varsayilan_tarih_iso()
    yanit = client.get("/gecmis/gonderilenler")
    assert yanit.status_code == 200
    assert f'id="personalTarihBaslangic" value="{bas}"' in yanit.text
    assert f'id="personalTarihBitis" value="{bit}"' in yanit.text
    assert "Kisisel Gonderilen" in yanit.text


def test_onay_kuyrugu_gecersiz_sayfa(client, veritabani):
    app.dependency_overrides[aktif_kullanici] = lambda: _YONETICI
    yanit = client.get("/onay-kuyrugu?sayfa=0")
    assert yanit.status_code == 200
    yanit2 = client.get("/onay-kuyrugu?birim=yok-boyle-birim")
    assert yanit2.status_code == 200


def _assert_departman_arama_kutusu(html: str, *, secili: str | None = None) -> None:
    assert 'name="birim"' in html
    assert "data-departman-combo" in html
    assert 'role="combobox"' in html
    assert 'role="listbox"' in html
    assert "Tüm departmanlar" in html
    assert ">Finans<" in html
    assert 'data-value="finans"' in html
    assert 'data-value="it-yazilim"' in html
    assert 'data-label="IT Yazilim"' in html
    if secili:
        assert f'option value="{secili}" selected' in html
        assert f'data-value="{secili}"' in html


def test_raporlar_departman_acilir_liste_ve_filtre(client, veritabani):
    app.dependency_overrides[aktif_kullanici] = lambda: _YONETICI
    with Session(veritabani) as oturum:
        _onayli(oturum, "Finans Rapor")
        h = _onayli(oturum, "IT Rapor")
        h.olusturan_departman = "it"
        oturum.add(h)
        oturum.commit()

    yanit = client.get("/raporlar")
    assert yanit.status_code == 200
    _assert_departman_arama_kutusu(yanit.text)
    assert 'placeholder="Unit"' not in yanit.text
    assert "Departman" in yanit.text
    assert "Finans Rapor" in yanit.text
    assert "IT Rapor" in yanit.text

    filtrelenmis = client.get("/raporlar?birim=finans")
    assert filtrelenmis.status_code == 200
    assert "Finans Rapor" in filtrelenmis.text
    assert "IT Rapor" not in filtrelenmis.text
    _assert_departman_arama_kutusu(filtrelenmis.text, secili="finans")


def test_onay_kuyrugu_departman_acilir_liste(client, veritabani):
    app.dependency_overrides[aktif_kullanici] = lambda: _YONETICI
    yanit = client.get("/onay-kuyrugu")
    assert yanit.status_code == 200
    _assert_departman_arama_kutusu(yanit.text)


def test_gecmis_arama_departman_arama_kutusu_ve_filtre(client, veritabani):
    app.dependency_overrides[aktif_kullanici] = lambda: _YONETICI
    with Session(veritabani) as oturum:
        a = _onayli(oturum, "Finans Arama")
        a.olusturan_manager_zinciri = ["onur.simsek"]
        b = _onayli(oturum, "IT Arama")
        b.olusturan_departman = "it"
        b.olusturan_manager_zinciri = ["onur.simsek"]
        oturum.add(a)
        oturum.add(b)
        oturum.commit()

    yanit = client.get("/gecmis/arama")
    assert yanit.status_code == 200
    _assert_departman_arama_kutusu(yanit.text)

    filtrelenmis = client.get("/gecmis/arama?birim=finans")
    assert filtrelenmis.status_code == 200
    assert "Finans Arama" in filtrelenmis.text
    assert "IT Arama" not in filtrelenmis.text
    _assert_departman_arama_kutusu(filtrelenmis.text, secili="finans")
