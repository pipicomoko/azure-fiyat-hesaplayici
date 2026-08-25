import io

import pytest
from openpyxl import Workbook, load_workbook

from app.disa_aktar import (
    CokluExcelPaket,
    TahminBosHatasi,
    benzersiz_excel_sayfa_adi,
    calisma_kitabi_olustur,
    coklu_calisma_kitabi_olustur,
    excel_sayfa_adi,
)
from app.products.base import DisaAktarimSatiri


def test_bos_tahmin_disa_aktarilamaz():
    with pytest.raises(TahminBosHatasi):
        calisma_kitabi_olustur([], 0.0, "USD", "tr")


def test_excel_sayfa_adi_yasak_karakterleri_temizler():
    assert excel_sayfa_adi("2026 Q1/Q2 Planı") == "2026 Q1_Q2 Planı"
    assert "/" not in excel_sayfa_adi("a/b\\c*d?e:f[g]h")
    assert len(excel_sayfa_adi("x" * 50)) == 31
    ws = Workbook().active
    ws.title = excel_sayfa_adi("2026 Q1/Q2 Planı")  # ValueError olmamali


def test_benzersiz_excel_sayfa_adi_carpisinca_numaralar():
    mevcut = {"2026 Q1_Q2 Planı"}
    ikinci = benzersiz_excel_sayfa_adi("2026 Q1/Q2 Planı", mevcut)
    assert ikinci == "2026 Q1_Q2 Planı_2"
    assert len(ikinci) <= 31


def test_calisma_kitabi_basliklar_ve_toplam():
    satirlar = [
        DisaAktarimSatiri(
            urun="Managed Disks",
            yapilandirma_ozeti="Standard HDD - S4 - x1 - East US",
            bolge="East US",
            miktar=1,
            birim="disk",
            birim_fiyat=1.536,
            ara_toplam=1.536,
            servis_kategori="Storage",
        ),
        DisaAktarimSatiri(
            urun="Virtual Machines",
            yapilandirma_ozeti="1 x D2s v5 - Ubuntu - East US",
            bolge="East US",
            miktar=730,
            birim="saat",
            birim_fiyat=0.096,
            ara_toplam=70.08,
            servis_kategori="Compute",
        ),
    ]
    icerik = calisma_kitabi_olustur(satirlar, 71.616, "USD", "tr")

    kitap = load_workbook(io.BytesIO(icerik))
    sayfa = kitap.active
    satirlar_okunan = list(sayfa.iter_rows(values_only=True))

    assert satirlar_okunan[0][0] == "Microsoft Azure Estimate"
    assert satirlar_okunan[1][0] == "Your Estimate"
    baslik = satirlar_okunan[2]
    assert baslik[0] == "Service category"
    assert baslik[5] == "Estimated monthly cost"
    assert baslik[6] == "Indirim Yuzdesi"
    assert baslik[7] == "Indirimli Aylik Maliyet"
    assert baslik[8] == "İndirimli Yıllık Maliyeti"
    assert baslik[9] == "Estimated upfront cost"
    assert baslik[10] == "Yillik Tahmini Maliyet"
    assert satirlar_okunan[3][1] == "Managed Disks"
    assert satirlar_okunan[4][1] == "Virtual Machines"
    toplam_satiri = next(s for s in satirlar_okunan if s[3] == "Total")
    assert toplam_satiri[5] == 71.62
    assert toplam_satiri[8] == 0
    assert toplam_satiri[10] == round(71.616 * 12, 2)

    # Indirimsiz satirlar: tire, $ yok. Maliyet hucrelerinde $ formatı.
    disk_satir_no = 4
    disk = list(sayfa.iter_rows(min_row=disk_satir_no, max_row=disk_satir_no))[0]
    assert disk[6].value == "—"
    assert disk[7].value == "—"
    assert disk[8].value == "—"
    assert "$" not in (disk[6].number_format or "")
    assert "$" not in (disk[7].number_format or "")
    assert "$" in (disk[5].number_format or "")
    assert "$" in (disk[9].number_format or "")
    assert "$" in (disk[10].number_format or "")
    toplam_hucreler = next(
        row for row in sayfa.iter_rows() if row[3].value == "Total"
    )
    assert "$" in (toplam_hucreler[5].number_format or "")
    assert "$" in (toplam_hucreler[8].number_format or "")
    assert "$" in (toplam_hucreler[10].number_format or "")
    assert "$" not in (disk[1].number_format or "")  # urun adi


def _ornek_satir(ad: str = "Managed Disks") -> DisaAktarimSatiri:
    return DisaAktarimSatiri(
        urun=ad,
        yapilandirma_ozeti="S4",
        bolge="East US",
        miktar=1,
        birim="disk",
        birim_fiyat=1.5,
        ara_toplam=1.5,
        servis_kategori="Storage",
    )


def test_coklu_calisma_kitabi_tek_workbook_cok_sayfa():
    paketler = [
        CokluExcelPaket(
            ad="2026 Q1/Q2 Planı",
            kayit_id=1,
            satirlar=[_ornek_satir()],
            genel_toplam=1.5,
            para_birimi="USD",
            dil="tr",
        ),
        CokluExcelPaket(
            ad="Ikinci",
            kayit_id=2,
            satirlar=[_ornek_satir("VM")],
            genel_toplam=2.0,
            para_birimi="USD",
            dil="tr",
        ),
    ]
    icerik = coklu_calisma_kitabi_olustur(paketler)
    kitap = load_workbook(io.BytesIO(icerik))
    assert len(kitap.sheetnames) == 2
    assert "Q1_Q2" in kitap.sheetnames[0] or kitap.sheetnames[0].startswith("2026")
    assert "/" not in kitap.sheetnames[0]
    assert "Ikinci" in kitap.sheetnames


def test_coklu_calisma_kitabi_yuzlerce_kayit_makul_surede():
    """BUG-06: load_workbook+hucre kopya dongusu yerine dogrudan yazim."""
    import time

    paketler = [
        CokluExcelPaket(
            ad=f"Senaryo {i}",
            kayit_id=i,
            satirlar=[_ornek_satir()],
            genel_toplam=1.5,
            para_birimi="USD",
            dil="tr",
        )
        for i in range(200)
    ]
    basla = time.perf_counter()
    icerik = coklu_calisma_kitabi_olustur(paketler)
    sure = time.perf_counter() - basla
    kitap = load_workbook(io.BytesIO(icerik))
    assert len(kitap.sheetnames) == 200
    assert len(icerik) > 1000
    # Eski yol ~2000 kayitta ~48sn; 200 kayit dogrudan yazimda saniyeler olmali
    assert sure < 15.0, f"coklu excel beklenenden yavas: {sure:.2f}s"


def test_indirimli_yillik_sutunu_formulu_ve_toplam():
    """Indirimli yillik = indirimli_aylik * 12. Yillik Tahmini Maliyet liste*12."""
    from app.disa_aktar import satirlara_indirim_uygula

    aylik = 100.0
    indirim = 10.0
    satir = DisaAktarimSatiri(
        urun="Virtual Machines",
        yapilandirma_ozeti="D2s v5",
        bolge="East US",
        miktar=1,
        birim="saat",
        birim_fiyat=aylik,
        ara_toplam=aylik,
        servis_kategori="Compute",
    )
    katilan = satirlara_indirim_uygula([satir], aylik, indirim)
    assert katilan == 90.0
    assert satir.indirimli_aylik == 90.0

    icerik = calisma_kitabi_olustur([satir], katilan, "USD", "tr")
    kitap = load_workbook(io.BytesIO(icerik))
    sayfa = kitap.active
    baslik = next(
        row for row in sayfa.iter_rows(values_only=True) if row[0] == "Service category"
    )
    assert baslik[8] == "İndirimli Yıllık Maliyeti"

    veri = next(row for row in sayfa.iter_rows() if row[1].value == "Virtual Machines")
    assert veri[5].value == 100.0
    assert veri[6].value == 10.0
    assert veri[7].value == 90.0
    assert veri[8].value == 1080.0  # indirimli yillik = 90 * 12
    assert veri[10].value == 1200.0  # Yillik Tahmini Maliyet = liste 100 * 12
    assert "$" in (veri[5].number_format or "")
    assert "$" in (veri[7].number_format or "")
    assert "$" in (veri[8].number_format or "")
    assert "$" not in (veri[6].number_format or "")  # yuzde

    toplam = next(row for row in sayfa.iter_rows() if row[3].value == "Total")
    assert toplam[5].value == 90.0  # aylik toplam indirimli kalir
    assert toplam[8].value == 1080.0
    assert toplam[10].value == 1200.0  # yillik toplam liste * 12
    assert "$" in (toplam[8].number_format or "")


def test_indirim_yoksa_indirimli_yillik_tire_kalir():
    satir = _ornek_satir()
    icerik = calisma_kitabi_olustur([satir], 1.5, "USD", "tr")
    kitap = load_workbook(io.BytesIO(icerik))
    sayfa = kitap.active
    veri = next(row for row in sayfa.iter_rows() if row[1].value == "Managed Disks")
    assert veri[6].value == "—"
    assert veri[7].value == "—"
    assert veri[8].value == "—"
    assert "$" not in (veri[8].number_format or "")
    assert veri[10].value == round(1.5 * 12, 2)
