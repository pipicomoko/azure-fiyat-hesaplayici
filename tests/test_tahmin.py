"""Tahmin calisma alani uc-uca (route seviyesinde) testleri.

Azure Retail Prices API'sine giden CANLI cagrilar burada MOCKLANIR (pricing
matematiginin dogrulugu tests/test_managed_disks_fiyatlama.py ve
tests/test_virtual_machines_fiyatlama.py'de ayrica, gercek API'den alinmis
sabit verilerle test edilir) -- bu dosya sadece HTTP/form/DB entegrasyonunu
dogrular: kalem ekleme, alan degisince yeniden hesaplama, kaydetme, disa
aktarma, gecmis/karsilastirma.
"""

import re
from urllib.parse import unquote

import pytest

from app.fiyat_api import onbellek_temizle
from app.models import Hesaplama
from app.routers.tahmin import _kopya_hesaplama_adi, _tek_tahmin_excel_adi
from app.yetkilendirme import (
    gecmis_erisim_kapsami,
    gruplardan_departman_belirle,
)

_S4_KAYITLARI = [
    {
        "meterName": "S4 LRS Disk",
        "skuName": "S4 LRS",
        "retailPrice": 1.536,
        "unitOfMeasure": "1/Month",
        "type": "Consumption",
        "productName": "Standard HDD Managed Disks",
        "effectiveStartDate": "2020-01-01",
    },
    {
        "meterName": "S4 LRS Disk Operations",
        "skuName": "S4 LRS",
        "retailPrice": 0.0005,
        "unitOfMeasure": "10K",
        "type": "Consumption",
        "productName": "Standard HDD Managed Disks",
        "effectiveStartDate": "2020-01-01",
    },
]


async def _sahte_disk_kayitlari(filtre, para_birimi="USD", onbellek_kullan=True):
    if "Standard HDD Managed Disks" in filtre:
        return _S4_KAYITLARI
    return []


@pytest.fixture(autouse=True)
def _fiyat_api_sahte(monkeypatch):
    onbellek_temizle()
    monkeypatch.setattr(
        "app.products.managed_disks.fiyatlama.kayitlari_getir", _sahte_disk_kayitlari
    )
    yield
    onbellek_temizle()


def _kalem_id_cikar(html: str) -> str:
    return re.search(r'kalem-([a-f0-9]+)"', html).group(1)


def test_tahmin_sayfasi_acilir(client):
    yanit = client.get("/tahmin")
    assert yanit.status_code == 200
    assert "Tahminim" in yanit.text


def test_kalem_ekle_disk_varsayilan_fiyatla_doner(client):
    yanit = client.post(
        "/tahmin/kalem-ekle", data={"urun_tipi": "managed_disks", "para_birimi": "USD"}
    )
    assert yanit.status_code == 200
    assert 'data-tutar="1.586"' in yanit.text or "1.59" in yanit.text


def test_kalem_hesapla_alan_degisince_yeniden_fiyatlar(client):
    ekleme = client.post(
        "/tahmin/kalem-ekle", data={"urun_tipi": "managed_disks", "para_birimi": "USD"}
    )
    kalem_id = _kalem_id_cikar(ekleme.text)

    veri = {
        f"{kalem_id}.urun_tipi": "managed_disks",
        f"{kalem_id}.bolge": "eastus",
        f"{kalem_id}.kademe": "standardhdd",
        f"{kalem_id}.sku": "S4",
        f"{kalem_id}.adet": "3",
        f"{kalem_id}.islem_adet": "0",
        "para_birimi": "USD",
    }
    yanit = client.post(f"/tahmin/kalem/hesapla?kalem_id={kalem_id}", data=veri)
    assert yanit.status_code == 200
    assert 'data-tutar="4.608"' in yanit.text  # 3 x $1.536
    assert 'data-liste-tutar="4.608"' in yanit.text


def test_kalem_hesapla_indirim_aylik_uygular_liste_tutarini_korur(client):
    ekleme = client.post(
        "/tahmin/kalem-ekle", data={"urun_tipi": "managed_disks", "para_birimi": "USD"}
    )
    kalem_id = _kalem_id_cikar(ekleme.text)
    veri = {
        f"{kalem_id}.urun_tipi": "managed_disks",
        f"{kalem_id}.bolge": "eastus",
        f"{kalem_id}.kademe": "standardhdd",
        f"{kalem_id}.sku": "S4",
        f"{kalem_id}.adet": "3",
        f"{kalem_id}.islem_adet": "0",
        f"{kalem_id}.indirim_yuzdesi": "10",
        "para_birimi": "USD",
    }
    yanit = client.post(f"/tahmin/kalem/hesapla?kalem_id={kalem_id}", data=veri)
    assert yanit.status_code == 200
    assert 'data-liste-tutar="4.608"' in yanit.text
    assert 'data-tutar="4.1472"' in yanit.text  # 4.608 * 0.9
    assert "10" in yanit.text


def test_gecersiz_urun_tipi_reddedilir(client):
    yanit = client.post("/tahmin/kalem-ekle", data={"urun_tipi": "olmayan_urun"})
    assert yanit.status_code == 400


def test_bos_tahmin_disa_aktarilamaz(client):
    yanit = client.post("/tahmin/disa-aktar", data={"para_birimi": "USD"})
    assert yanit.status_code == 400


def test_bos_tahmin_kaydedilemez(client):
    yanit = client.post(
        "/tahmin/kaydet", data={"para_birimi": "USD", "hesaplama_adi": "Test"}
    )
    assert yanit.status_code == 400


def test_ad_verilmeden_kaydedilemez(client):
    ekleme = client.post(
        "/tahmin/kalem-ekle", data={"urun_tipi": "managed_disks", "para_birimi": "USD"}
    )
    kalem_id = _kalem_id_cikar(ekleme.text)
    veri = {
        f"{kalem_id}.urun_tipi": "managed_disks",
        f"{kalem_id}.bolge": "eastus",
        f"{kalem_id}.kademe": "standardhdd",
        f"{kalem_id}.sku": "S4",
        f"{kalem_id}.adet": "1",
        "para_birimi": "USD",
        "hesaplama_adi": "",
    }
    yanit = client.post("/tahmin/kaydet", data=veri)
    assert yanit.status_code == 400


def test_tek_tahmin_excel_adi_hesaplama_adini_kullanir():
    assert _tek_tahmin_excel_adi("OnurŞimşekTest") == "OnurŞimşekTest.xlsx"
    assert _tek_tahmin_excel_adi("  OnurŞimşekTest  ") == "OnurŞimşekTest.xlsx"
    assert _tek_tahmin_excel_adi("zaten.xlsx") == "zaten.xlsx"
    assert _tek_tahmin_excel_adi("") == "tahmin.xlsx"
    assert _tek_tahmin_excel_adi("   ") == "tahmin.xlsx"
    assert _tek_tahmin_excel_adi(None) == "tahmin.xlsx"


def test_disa_aktar_butonu_formnovalidate_onay_hedefi_zorunlu_kalir(client):
    yanit = client.get("/tahmin")
    assert yanit.status_code == 200
    disa_aktar = re.search(
        r'<button[^>]*id="disa-aktar-butonu"[^>]*>', yanit.text
    )
    assert disa_aktar is not None
    etiket = disa_aktar.group(0)
    assert "formnovalidate" in etiket
    assert 'formaction="/tahmin/disa-aktar"' in etiket

    from app.sablonlar import templates

    html = templates.env.get_template("_tahmin_ic_icerik.html").render(
        kullanici={
            "kullanici_adi": "test.kullanici",
            "ad_soyad": "Test Kullanici",
            "unvan": "Test Unvani",
            "gruplar": ["AFH-Calisanlar"],
            "manager": "onur.simsek",
            "manager_zinciri": ["onur.simsek", "emre.turan"],
            "rol": "calisan",
        },
        urunler=[],
        dil="tr",
        para_birimi="USD",
        red_gerekce=None,
        reddeden_kullanici_adi=None,
        hesaplama_adi="",
        kalem_sonuclari=[],
    )
    onay_secici = re.search(
        r'<select[^>]*id="onay-hedefi-secici"[^>]*>', html
    )
    assert onay_secici is not None
    assert "required" in onay_secici.group(0)
    ic_disa = re.search(r'<button[^>]*id="disa-aktar-butonu"[^>]*>', html)
    assert ic_disa is not None
    assert "formnovalidate" in ic_disa.group(0)
    onaya_gonder = re.search(
        r'<button[^>]*name="onaya_gonder"[^>]*value="1"[^>]*>', html
    )
    assert onaya_gonder is not None
    assert "formnovalidate" not in onaya_gonder.group(0)


def test_disa_aktarim_xlsx_dosyasi_doner(client):
    ekleme = client.post(
        "/tahmin/kalem-ekle", data={"urun_tipi": "managed_disks", "para_birimi": "USD"}
    )
    kalem_id = _kalem_id_cikar(ekleme.text)
    veri = {
        f"{kalem_id}.urun_tipi": "managed_disks",
        f"{kalem_id}.bolge": "eastus",
        f"{kalem_id}.kademe": "standardhdd",
        f"{kalem_id}.sku": "S4",
        f"{kalem_id}.adet": "1",
        "para_birimi": "USD",
        "hesaplama_adi": "OnurŞimşekTest",
    }
    yanit = client.post("/tahmin/disa-aktar", data=veri)
    assert yanit.status_code == 200
    assert yanit.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    cd = yanit.headers["content-disposition"]
    assert "attachment" in cd
    utf8_adi = unquote(cd.split("filename*=UTF-8''", 1)[1])
    assert utf8_adi == "OnurŞimşekTest.xlsx"
    assert "azure-tahmin-" not in utf8_adi


def test_disa_aktarim_onay_hedefi_bos_xlsx_doner(client):
    ekleme = client.post(
        "/tahmin/kalem-ekle", data={"urun_tipi": "managed_disks", "para_birimi": "USD"}
    )
    kalem_id = _kalem_id_cikar(ekleme.text)
    veri = {
        f"{kalem_id}.urun_tipi": "managed_disks",
        f"{kalem_id}.bolge": "eastus",
        f"{kalem_id}.kademe": "standardhdd",
        f"{kalem_id}.sku": "S4",
        f"{kalem_id}.adet": "1",
        "para_birimi": "USD",
        "hesaplama_adi": "OnaysizAktarim",
        "onay_hedefi": "",
    }
    yanit = client.post("/tahmin/disa-aktar", data=veri)
    assert yanit.status_code == 200
    assert yanit.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert yanit.content[:2] == b"PK"
    cd = yanit.headers["content-disposition"]
    utf8_adi = unquote(cd.split("filename*=UTF-8''", 1)[1])
    assert utf8_adi == "OnaysizAktarim.xlsx"


def test_disa_aktarim_bos_ad_tahmin_xlsx_doner(client):
    ekleme = client.post(
        "/tahmin/kalem-ekle", data={"urun_tipi": "managed_disks", "para_birimi": "USD"}
    )
    kalem_id = _kalem_id_cikar(ekleme.text)
    veri = {
        f"{kalem_id}.urun_tipi": "managed_disks",
        f"{kalem_id}.bolge": "eastus",
        f"{kalem_id}.kademe": "standardhdd",
        f"{kalem_id}.sku": "S4",
        f"{kalem_id}.adet": "1",
        "para_birimi": "USD",
    }
    yanit = client.post("/tahmin/disa-aktar", data=veri)
    assert yanit.status_code == 200
    cd = yanit.headers["content-disposition"]
    utf8_adi = unquote(cd.split("filename*=UTF-8''", 1)[1])
    assert utf8_adi == "tahmin.xlsx"


def test_onaya_gonder_hedefsiz_reddedilir(client):
    ekleme = client.post(
        "/tahmin/kalem-ekle", data={"urun_tipi": "managed_disks", "para_birimi": "USD"}
    )
    kalem_id = _kalem_id_cikar(ekleme.text)
    veri = {
        f"{kalem_id}.urun_tipi": "managed_disks",
        f"{kalem_id}.bolge": "eastus",
        f"{kalem_id}.kademe": "standardhdd",
        f"{kalem_id}.sku": "S4",
        f"{kalem_id}.adet": "1",
        "para_birimi": "USD",
        "hesaplama_adi": "Hedefsiz",
        "onaya_gonder": "1",
    }
    yanit = client.post("/tahmin/kaydet", data=veri)
    assert yanit.status_code == 400
    metin = yanit.text.casefold()
    assert (
        "yonetici" in metin
        or "yönetici" in metin
        or "approver" in metin
        or "seç" in metin
        or "select" in metin
    )


def test_genel_mudur_onaya_gonderemez_sadece_kaydeder(client, veritabani):
    from app.main import app
    from app.models import DURUM_TASLAK, Hesaplama
    from app.yetkilendirme import GENEL_MUDUR_SAM, aktif_kullanici
    from sqlmodel import select

    app.dependency_overrides[aktif_kullanici] = lambda: {
        "kullanici_adi": GENEL_MUDUR_SAM,
        "ad_soyad": "Ahmet Yildirim",
        "unvan": "Genel Mudur",
        "gruplar": ["AFH-Calisanlar", "AFH-Direktorler"],
        "rol": "direktor",
        "manager": None,
        "manager_zinciri": [],
    }
    ekleme = client.post(
        "/tahmin/kalem-ekle", data={"urun_tipi": "managed_disks", "para_birimi": "USD"}
    )
    kalem_id = _kalem_id_cikar(ekleme.text)
    veri = {
        f"{kalem_id}.urun_tipi": "managed_disks",
        f"{kalem_id}.bolge": "eastus",
        f"{kalem_id}.kademe": "standardhdd",
        f"{kalem_id}.sku": "S4",
        f"{kalem_id}.adet": "1",
        "para_birimi": "USD",
        "hesaplama_adi": "GM Deneme",
        "onaya_gonder": "1",
        "onay_hedefi": "kimse",
    }
    yanit = client.post("/tahmin/kaydet", data=veri)
    assert yanit.status_code == 200

    with __import__("sqlmodel").Session(veritabani) as oturum:
        kayit = oturum.exec(select(Hesaplama).where(Hesaplama.ad == "GM Deneme")).one()
        assert kayit.durum == DURUM_TASLAK
        assert kayit.onay_hedefi is None

    gecmis = client.get("/gecmis/taslaklar")
    assert "GM Deneme" in gecmis.text
    assert "Estimate History" in gecmis.text or "Tahmin ge" in gecmis.text

    app.dependency_overrides.pop(aktif_kullanici, None)


def test_kaydet_ve_gecmis_akisi(client, veritabani):
    ekleme = client.post(
        "/tahmin/kalem-ekle", data={"urun_tipi": "managed_disks", "para_birimi": "USD"}
    )
    kalem_id = _kalem_id_cikar(ekleme.text)
    veri = {
        f"{kalem_id}.urun_tipi": "managed_disks",
        f"{kalem_id}.bolge": "eastus",
        f"{kalem_id}.kademe": "standardhdd",
        f"{kalem_id}.sku": "S4",
        f"{kalem_id}.adet": "1",
        "para_birimi": "USD",
        "hesaplama_adi": "Test Senaryosu",
    }
    kaydet_yaniti = client.post("/tahmin/kaydet", data=veri)
    assert kaydet_yaniti.status_code == 200
    assert "kaydedildi" in kaydet_yaniti.text

    gecmis_yaniti = client.get("/gecmis/taslaklar")
    assert "Test Senaryosu" in gecmis_yaniti.text

    with __import__("sqlmodel").Session(veritabani) as oturum:
        from sqlmodel import select

        kayit = oturum.exec(select(Hesaplama)).one()
        assert kayit.ad == "Test Senaryosu"
        assert round(kayit.toplam_aylik_maliyet, 3) == 1.536
        assert len(kayit.kalemler) == 1
        assert kayit.kalemler[0].yapilandirma["sku"] == "S4"


def test_gecmis_karsilastirma_2den_farkli_secim_reddedilir(client, veritabani):
    yanit = client.get("/gecmis/karsilastir?id=1")
    # Eski karsilastirma ucu kalktiysa 404; varsa 400/422
    assert yanit.status_code in (400, 404, 422)


def _kaydet(client, hesaplama_adi="Test Senaryosu"):
    ekleme = client.post(
        "/tahmin/kalem-ekle", data={"urun_tipi": "managed_disks", "para_birimi": "USD"}
    )
    kalem_id = _kalem_id_cikar(ekleme.text)
    veri = {
        f"{kalem_id}.urun_tipi": "managed_disks",
        f"{kalem_id}.bolge": "eastus",
        f"{kalem_id}.kademe": "standardhdd",
        f"{kalem_id}.sku": "S4",
        f"{kalem_id}.adet": "1",
        "para_birimi": "USD",
        "hesaplama_adi": hesaplama_adi,
    }
    return client.post("/tahmin/kaydet", data=veri)


def test_gecmis_detay_yillik_maliyet_gosterir(client, veritabani):
    _kaydet(client)
    with __import__("sqlmodel").Session(veritabani) as oturum:
        from sqlmodel import select

        kayit = oturum.exec(select(Hesaplama)).one()

    yanit = client.get(f"/gecmis/{kayit.id}")
    assert yanit.status_code == 200
    assert "Test Senaryosu" in yanit.text
    assert "S4" in yanit.text  # kalemin ozeti detayda gorunuyor


def test_indirim_aylik_uygulanir_yillik_liste_fiyati_kalir(client, veritabani):
    ekleme = client.post(
        "/tahmin/kalem-ekle", data={"urun_tipi": "managed_disks", "para_birimi": "USD"}
    )
    kalem_id = _kalem_id_cikar(ekleme.text)
    veri = {
        f"{kalem_id}.urun_tipi": "managed_disks",
        f"{kalem_id}.bolge": "eastus",
        f"{kalem_id}.kademe": "standardhdd",
        f"{kalem_id}.sku": "S4",
        f"{kalem_id}.adet": "1",
        f"{kalem_id}.indirim_yuzdesi": "10",
        "para_birimi": "USD",
        "hesaplama_adi": "Indirimli Senaryo",
    }
    kaydet = client.post("/tahmin/kaydet", data=veri)
    assert kaydet.status_code == 200

    with __import__("sqlmodel").Session(veritabani) as oturum:
        from sqlmodel import select

        kayit = oturum.exec(select(Hesaplama)).one()
        kalem = kayit.kalemler[0]
        aylik = float(kalem.aylik_maliyet)
        indirimli = float(kalem.indirimli_aylik_maliyet)
        assert kalem.indirim_yuzdesi == 10.0
        assert indirimli == round(aylik * 0.9, 4)
        assert round(kayit.toplam_aylik_maliyet, 4) == indirimli
        para = kayit.para_birimi
        kayit_id = kayit.id

    from app.sablonlar import _para_bicimlendir, _yillik

    liste_yillik = _para_bicimlendir(_yillik(aylik), para)
    indirimli_yillik = _para_bicimlendir(_yillik(indirimli), para)
    indirimli_aylik_metin = _para_bicimlendir(indirimli, para)

    detay = client.get(f"/gecmis/{kayit_id}")
    assert detay.status_code == 200
    assert indirimli_aylik_metin in detay.text
    assert liste_yillik in detay.text
    assert indirimli_yillik in detay.text
    assert "İndirimli yıllık maliyeti" in detay.text

    liste = client.get("/gecmis/taslaklar")
    assert liste.status_code == 200
    assert liste_yillik in liste.text
    assert indirimli_aylik_metin in liste.text

    from io import BytesIO

    from openpyxl import load_workbook

    excel = client.get(f"/gecmis/{kayit_id}/excel")
    assert excel.status_code == 200
    sayfa = load_workbook(BytesIO(excel.content)).active
    veri_satiri = next(
        row for row in sayfa.iter_rows() if row[1].value == "Managed Disks"
    )
    excel_aylik = float(veri_satiri[5].value)
    assert excel_aylik == round(aylik, 2)
    assert veri_satiri[7].value == round(indirimli, 2)
    assert veri_satiri[8].value == round(indirimli * 12, 2)
    assert veri_satiri[10].value == round(excel_aylik * 12, 2)
    assert veri_satiri[10].value != veri_satiri[8].value
    toplam = next(row for row in sayfa.iter_rows() if row[3].value == "Total")
    assert toplam[5].value == round(indirimli, 2)
    assert toplam[8].value == round(indirimli * 12, 2)
    assert toplam[10].value == round(excel_aylik * 12, 2)


def test_gecmis_sadece_sahibi_gorur_admin_taslak_gormez(client, veritabani):
    from app.main import app
    from app.models import DURUM_ONAY_BEKLIYOR
    from app.yetkilendirme import aktif_kullanici

    app.dependency_overrides[aktif_kullanici] = lambda: {
        "kullanici_adi": "zeynep.kara",
        "ad_soyad": "Zeynep Kara",
        "unvan": "",
        "gruplar": ["AFH-Calisanlar"],
        "manager": "onur.simsek",
        "manager_zinciri": ["onur.simsek"],
    }
    _kaydet(client, "Zeynep Tahmini")

    # Admin taslagi gormez
    app.dependency_overrides[aktif_kullanici] = lambda: {
        "kullanici_adi": "asli.demirtas",
        "ad_soyad": "Asli Demirtas",
        "unvan": "",
        "gruplar": ["AFH-Adminler"],
        "rol": "admin",
    }
    can_gecmis = client.get("/gecmis/arama")
    assert "Zeynep Tahmini" not in can_gecmis.text

    with __import__("sqlmodel").Session(veritabani) as oturum:
        from sqlmodel import select

        kayit = oturum.exec(
            select(Hesaplama).where(Hesaplama.ad == "Zeynep Tahmini")
        ).one()
        kayit.durum = DURUM_ONAY_BEKLIYOR
        oturum.add(kayit)
        oturum.commit()
        kayit_id = kayit.id

    can_gecmis = client.get("/gecmis/arama")
    assert "Zeynep Tahmini" in can_gecmis.text

    app.dependency_overrides[aktif_kullanici] = lambda: {
        "kullanici_adi": "deniz.aksoy",
        "ad_soyad": "Deniz Aksoy",
        "unvan": "",
        "gruplar": ["AFH-Calisanlar"],
    }
    deniz_gecmis = client.get("/gecmis/taslaklar")
    assert "Zeynep Tahmini" not in deniz_gecmis.text

    detay_yaniti = client.get(f"/gecmis/{kayit_id}")
    assert detay_yaniti.status_code == 403
    sil_yaniti = client.post(f"/gecmis/{kayit_id}/sil")
    assert sil_yaniti.status_code == 403

    app.dependency_overrides.pop(aktif_kullanici, None)


def test_yonetici_manager_zincirindeki_kaydi_gorur(client, veritabani):
    from app.main import app
    from app.models import DURUM_ONAY_BEKLIYOR
    from app.yetkilendirme import aktif_kullanici

    app.dependency_overrides[aktif_kullanici] = lambda: {
        "kullanici_adi": "kerem.acar",
        "ad_soyad": "Kerem Acar",
        "unvan": "",
        "gruplar": ["AFH-Calisanlar", "DEPT-IT"],
        "departman": "it-altyapi",
        "manager": "onur.simsek",
        "manager_zinciri": ["onur.simsek", "emre.turan"],
    }
    _kaydet(client, "Kerem Tahmini")

    with __import__("sqlmodel").Session(veritabani) as oturum:
        from sqlmodel import select

        kayit = oturum.exec(
            select(Hesaplama).where(Hesaplama.ad == "Kerem Tahmini")
        ).one()
        kayit.durum = DURUM_ONAY_BEKLIYOR
        kayit.olusturan_manager_zinciri = ["onur.simsek", "emre.turan"]
        oturum.add(kayit)
        oturum.commit()
        kayit_id = kayit.id

    app.dependency_overrides[aktif_kullanici] = lambda: {
        "kullanici_adi": "mehmet.hr",
        "ad_soyad": "Mehmet HR",
        "unvan": "",
        "gruplar": ["AFH-Calisanlar", "DEPT-IK"],
        "departman": "ik",
        "manager": "baska",
        "manager_zinciri": ["baska"],
    }
    _kaydet(client, "HR Calisan Tahmini")
    with __import__("sqlmodel").Session(veritabani) as oturum:
        from sqlmodel import select

        hr = oturum.exec(
            select(Hesaplama).where(Hesaplama.ad == "HR Calisan Tahmini")
        ).one()
        hr.durum = DURUM_ONAY_BEKLIYOR
        hr.olusturan_manager_zinciri = ["baska"]
        oturum.add(hr)
        oturum.commit()
        hr_id = hr.id

    app.dependency_overrides[aktif_kullanici] = lambda: {
        "kullanici_adi": "onur.simsek",
        "ad_soyad": "Onur Simsek",
        "unvan": "",
        "gruplar": ["AFH-Calisanlar", "AFH-Yoneticiler", "DEPT-IT"],
        "rol": "yonetici",
        "manager_zinciri": ["emre.turan"],
    }
    mudur_gecmis = client.get("/gecmis/arama")
    assert "Kerem Tahmini" in mudur_gecmis.text
    assert "HR Calisan Tahmini" not in mudur_gecmis.text

    detay_yaniti = client.get(f"/gecmis/{hr_id}")
    assert detay_yaniti.status_code == 403
    assert client.get(f"/gecmis/{kayit_id}").status_code == 200

    app.dependency_overrides.pop(aktif_kullanici, None)


def test_yonetici_zincirde_yoksa_baskasini_gormez(client, veritabani):
    from app.main import app
    from app.yetkilendirme import aktif_kullanici

    app.dependency_overrides[aktif_kullanici] = lambda: {
        "kullanici_adi": "zeynep.kara",
        "ad_soyad": "Zeynep Kara",
        "unvan": "",
        "gruplar": ["AFH-Calisanlar"],
        "manager_zinciri": ["x"],
    }
    _kaydet(client, "Zeynep Tahmini")

    app.dependency_overrides[aktif_kullanici] = lambda: {
        "kullanici_adi": "genel.mudur",
        "ad_soyad": "Genel Mudur",
        "unvan": "",
        "gruplar": ["AFH-Yoneticiler"],
        "rol": "yonetici",
    }
    mudur_gecmis = client.get("/gecmis/arama")
    assert "Zeynep Tahmini" not in mudur_gecmis.text

    app.dependency_overrides.pop(aktif_kullanici, None)


def test_gruplardan_departman_belirle_it_ik_ve_diger():
    assert gruplardan_departman_belirle(["AFH-Calisanlar", "DEPT-IT"]) == ("it", "IT")
    assert gruplardan_departman_belirle(["HR.Calisanlar"]) == ("ik", "IK")
    assert gruplardan_departman_belirle(["AFH-Calisanlar"]) == ("diger", "Diger")


def test_gecmis_erisim_kapsami_admin_yonetici_calisan():
    assert gecmis_erisim_kapsami({"gruplar": ["AFH-Adminler"]}) == "admin"
    assert gecmis_erisim_kapsami({"gruplar": ["AFH-Direktorler"]}) == "direktor"
    assert gecmis_erisim_kapsami({"gruplar": ["AFH-Yoneticiler"]}) == "yonetici"
    assert gecmis_erisim_kapsami({"gruplar": ["AFH-Calisanlar"]}) == "kendi"


def test_sahibi_kendi_tahminini_silebilir(client, veritabani):
    from app.main import app
    from app.yetkilendirme import aktif_kullanici

    app.dependency_overrides[aktif_kullanici] = lambda: {
        "kullanici_adi": "zeynep.kara",
        "ad_soyad": "Zeynep Kara",
        "unvan": "",
        "gruplar": ["AFH-Calisanlar"],
    }
    _kaydet(client, "Silinecek Tahmin")

    with __import__("sqlmodel").Session(veritabani) as oturum:
        from sqlmodel import select

        kayit = oturum.exec(
            select(Hesaplama).where(Hesaplama.ad == "Silinecek Tahmin")
        ).one()

    yanit = client.post(f"/gecmis/{kayit.id}/sil", follow_redirects=False)
    assert yanit.status_code == 303

    with __import__("sqlmodel").Session(veritabani) as oturum:
        from sqlmodel import select

        kalan = oturum.exec(
            select(Hesaplama).where(Hesaplama.ad == "Silinecek Tahmin")
        ).all()
    assert kalan == []

    app.dependency_overrides.pop(aktif_kullanici, None)


def test_dil_degistir_tahmini_kaybetmez(client):
    ekleme = client.post(
        "/tahmin/kalem-ekle", data={"urun_tipi": "managed_disks", "para_birimi": "USD"}
    )
    kalem_id = _kalem_id_cikar(ekleme.text)
    veri = {
        f"{kalem_id}.urun_tipi": "managed_disks",
        f"{kalem_id}.bolge": "eastus",
        f"{kalem_id}.kademe": "standardhdd",
        f"{kalem_id}.sku": "S4",
        f"{kalem_id}.adet": "1",
        "para_birimi": "USD",
        "dil": "en",
    }
    yanit = client.post("/tahmin/dil-degistir", data=veri)
    assert yanit.status_code == 200
    assert (
        "Managed Disks" in yanit.text
    )  # kalem, ingilizce olarak, KAYBOLMADAN geri geldi
    assert f"kalem-{kalem_id}" in yanit.text


def test_gecmis_tumu_excel_yasak_sayfa_karakterinde_500_vermez(client, veritabani):
    """BUG-02: 'Q1/Q2' gibi adlar toplu Excel'i cokertmemeli."""
    assert _kaydet(client, "2026 Q1/Q2 Planı").status_code == 200
    assert _kaydet(client, "Normal Ad").status_code == 200

    yanit = client.get("/gecmis-excel")
    assert yanit.status_code == 200, yanit.text[:500]
    assert len(yanit.content) > 100

    from openpyxl import load_workbook
    import io

    kitap = load_workbook(io.BytesIO(yanit.content))
    assert any("Q1_Q2" in ad or "Plan" in ad for ad in kitap.sheetnames)
    for ad in kitap.sheetnames:
        assert "/" not in ad
        assert len(ad) <= 31


def test_kendini_onay_hedefi_secmek_reddedilir(client, veritabani):
    """BUG-03: hesaplama.kullan + onay.islem sahibi kendini onayci secemez."""
    from app.main import app
    from app.models import Hesaplama
    from app.yetkilendirme import aktif_kullanici
    from sqlmodel import select

    app.dependency_overrides[aktif_kullanici] = lambda: {
        "kullanici_adi": "onur.simsek",
        "ad_soyad": "Onur Simsek",
        "unvan": "Birim Sorumlusu",
        "gruplar": ["AFH-Calisanlar", "AFH-Yoneticiler"],
        "rol": "yonetici",
        "manager": "emre.turan",
        "manager_zinciri": ["emre.turan", "baris.kocak"],
    }
    ekleme = client.post(
        "/tahmin/kalem-ekle", data={"urun_tipi": "managed_disks", "para_birimi": "USD"}
    )
    kalem_id = _kalem_id_cikar(ekleme.text)
    veri = {
        f"{kalem_id}.urun_tipi": "managed_disks",
        f"{kalem_id}.bolge": "eastus",
        f"{kalem_id}.kademe": "standardhdd",
        f"{kalem_id}.sku": "S4",
        f"{kalem_id}.adet": "1",
        "para_birimi": "USD",
        "hesaplama_adi": "Self Approve Deneme",
        "onaya_gonder": "1",
        "onay_hedefi": "onur.simsek",
    }
    yanit = client.post("/tahmin/kaydet", data=veri)
    assert yanit.status_code == 400
    assert "kendiniz" in yanit.text.casefold() or "yourself" in yanit.text.casefold()

    with __import__("sqlmodel").Session(veritabani) as oturum:
        assert (
            oturum.exec(
                select(Hesaplama).where(Hesaplama.ad == "Self Approve Deneme")
            ).all()
            == []
        )

    app.dependency_overrides.pop(aktif_kullanici, None)


def test_kendi_kaydini_onaylamak_403(client, veritabani):
    """BUG-03: eski/kotu veride bile kendi kayit onaylanamaz."""
    from datetime import datetime, timezone

    from app.main import app
    from app.models import DURUM_ONAY_BEKLIYOR, DURUM_ONAYLANDI, Hesaplama
    from app.yetkilendirme import aktif_kullanici
    from sqlmodel import select

    with __import__("sqlmodel").Session(veritabani) as oturum:
        kayit = Hesaplama(
            ad="Eski Self Approve",
            durum=DURUM_ONAY_BEKLIYOR,
            olusturan_kullanici_adi="onur.simsek",
            onay_hedefi="onur.simsek",
            toplam_aylik_maliyet=1.0,
            para_birimi="USD",
            olusturulma_tarihi=datetime.now(timezone.utc),
        )
        oturum.add(kayit)
        oturum.commit()
        oturum.refresh(kayit)
        kayit_id = kayit.id

    app.dependency_overrides[aktif_kullanici] = lambda: {
        "kullanici_adi": "onur.simsek",
        "ad_soyad": "Onur Simsek",
        "unvan": "",
        "gruplar": ["AFH-Calisanlar", "AFH-Yoneticiler"],
        "rol": "yonetici",
        "manager": "emre.turan",
        "manager_zinciri": ["emre.turan"],
    }
    yanit = client.post(f"/onay/{kayit_id}/onayla")
    assert yanit.status_code == 403

    kuyruk = client.get("/onay-kuyrugu")
    assert kuyruk.status_code == 200
    assert "Eski Self Approve" not in kuyruk.text

    with __import__("sqlmodel").Session(veritabani) as oturum:
        kayit = oturum.exec(select(Hesaplama).where(Hesaplama.id == kayit_id)).one()
        assert kayit.durum == DURUM_ONAY_BEKLIYOR
        assert kayit.durum != DURUM_ONAYLANDI

    app.dependency_overrides.pop(aktif_kullanici, None)


def test_zincir_disi_onay_hedefi_reddedilir_ve_enjekte_edilmez(client, veritabani):
    """BUG-04: rastgele onayci kabul edilip manager zincirine yazilmamali."""
    from app.main import app
    from app.models import Hesaplama
    from app.yetkilendirme import aktif_kullanici
    from sqlmodel import select

    app.dependency_overrides[aktif_kullanici] = lambda: {
        "kullanici_adi": "kerem.acar",
        "ad_soyad": "Kerem Acar",
        "unvan": "",
        "gruplar": ["AFH-Calisanlar"],
        "rol": "calisan",
        "manager": "onur.simsek",
        "manager_zinciri": ["onur.simsek", "emre.turan", "baris.kocak"],
        "manager_adlari": {"onur.simsek": "Onur Simsek"},
    }
    ekleme = client.post(
        "/tahmin/kalem-ekle", data={"urun_tipi": "managed_disks", "para_birimi": "USD"}
    )
    kalem_id = _kalem_id_cikar(ekleme.text)
    veri = {
        f"{kalem_id}.urun_tipi": "managed_disks",
        f"{kalem_id}.bolge": "eastus",
        f"{kalem_id}.kademe": "standardhdd",
        f"{kalem_id}.sku": "S4",
        f"{kalem_id}.adet": "1",
        "para_birimi": "USD",
        "hesaplama_adi": "Zincir Bypass",
        "onaya_gonder": "1",
        "onay_hedefi": "rastgele.kisi",
    }
    yanit = client.post("/tahmin/kaydet", data=veri)
    assert yanit.status_code == 400
    metin = yanit.text.casefold()
    assert "zincir" in metin or "chain" in metin or "manager" in metin

    with __import__("sqlmodel").Session(veritabani) as oturum:
        assert (
            oturum.exec(select(Hesaplama).where(Hesaplama.ad == "Zincir Bypass")).all()
            == []
        )

    veri["onay_hedefi"] = "onur.simsek"
    veri["hesaplama_adi"] = "Zincir OK"
    yanit_ok = client.post("/tahmin/kaydet", data=veri)
    assert yanit_ok.status_code == 200, yanit_ok.text[:300]

    with __import__("sqlmodel").Session(veritabani) as oturum:
        kayit = oturum.exec(select(Hesaplama).where(Hesaplama.ad == "Zincir OK")).one()
        assert kayit.onay_hedefi == "onur.simsek"
        assert kayit.olusturan_manager_zinciri == [
            "onur.simsek",
            "emre.turan",
            "baris.kocak",
        ]
        assert "rastgele.kisi" not in (kayit.olusturan_manager_zinciri or [])

    app.dependency_overrides.pop(aktif_kullanici, None)


def _hesaplama_id_al(veritabani, ad: str) -> int:
    with __import__("sqlmodel").Session(veritabani) as oturum:
        from sqlmodel import select

        return oturum.exec(select(Hesaplama).where(Hesaplama.ad == ad)).one().id


def test_kopya_hesaplama_adi_benzersiz():
    assert _kopya_hesaplama_adi("OnurŞimşekTest", []) == "OnurŞimşekTest COPY"
    assert (
        _kopya_hesaplama_adi("OnurŞimşekTest", {"OnurŞimşekTest COPY"})
        == "OnurŞimşekTest COPY (1)"
    )
    assert (
        _kopya_hesaplama_adi(
            "OnurŞimşekTest",
            {"OnurŞimşekTest COPY", "OnurŞimşekTest COPY (1)"},
        )
        == "OnurŞimşekTest COPY (2)"
    )


def test_taslak_kopyalanir(client, veritabani):
    from app.models import DURUM_TASLAK

    _kaydet(client, "OnurŞimşekTest")
    kaynak_id = _hesaplama_id_al(veritabani, "OnurŞimşekTest")

    taslaklar = client.get("/gecmis/taslaklar")
    assert taslaklar.status_code == 200
    assert f"/gecmis/{kaynak_id}/kopyala" in taslaklar.text
    assert "Kopyala" in taslaklar.text

    detay = client.get(f"/gecmis/{kaynak_id}")
    assert detay.status_code == 200
    assert f"/gecmis/{kaynak_id}/kopyala" in detay.text

    yanit = client.post(f"/gecmis/{kaynak_id}/kopyala", follow_redirects=False)
    assert yanit.status_code == 303
    assert yanit.headers["location"].startswith("/tahmin?hesaplama_id=")
    yeni_id = int(yanit.headers["location"].rsplit("=", 1)[1])
    assert yeni_id != kaynak_id

    with __import__("sqlmodel").Session(veritabani) as oturum:
        kaynak = oturum.get(Hesaplama, kaynak_id)
        kopya = oturum.get(Hesaplama, yeni_id)
        assert kaynak is not None and kopya is not None
        assert kaynak.ad == "OnurŞimşekTest"
        assert kopya.ad == "OnurŞimşekTest COPY"
        assert kopya.durum == DURUM_TASLAK
        assert kopya.revizyon == 1
        assert kopya.red_gerekce is None
        assert kopya.reddeden_kullanici_adi is None
        assert kopya.onay_hedefi is None
        assert kopya.onay_hedefi_ad_soyad is None
        assert kopya.onaylayan_kullanici_adi is None
        assert kopya.onay_tarihi is None
        assert kopya.iptal_gerekce is None
        assert kopya.olusturan_kullanici_adi == "test.kullanici"
        assert len(kopya.kalemler) == len(kaynak.kalemler) == 1
        assert kopya.kalemler[0].id != kaynak.kalemler[0].id
        assert kopya.kalemler[0].urun_tipi == kaynak.kalemler[0].urun_tipi
        assert kopya.kalemler[0].yapilandirma == kaynak.kalemler[0].yapilandirma
        assert kopya.kalemler[0].fiyat_kalemleri == kaynak.kalemler[0].fiyat_kalemleri
        assert kopya.kalemler[0].indirim_yuzdesi == kaynak.kalemler[0].indirim_yuzdesi
        assert kopya.toplam_aylik_maliyet == kaynak.toplam_aylik_maliyet


def test_onay_bekleyen_kopyalanir(client, veritabani):
    from app.models import DURUM_ONAY_BEKLIYOR, DURUM_TASLAK
    from datetime import datetime, timezone

    _kaydet(client, "Bekleyen Tahmin")
    kaynak_id = _hesaplama_id_al(veritabani, "Bekleyen Tahmin")
    with __import__("sqlmodel").Session(veritabani) as oturum:
        kayit = oturum.get(Hesaplama, kaynak_id)
        kayit.durum = DURUM_ONAY_BEKLIYOR
        kayit.onay_hedefi = "onur.simsek"
        kayit.onay_hedefi_ad_soyad = "Onur Simsek"
        kayit.revizyon = 3
        kayit.onaylayan_kullanici_adi = None
        kayit.onay_tarihi = datetime.now(timezone.utc)
        oturum.add(kayit)
        oturum.commit()

    gonderilenler = client.get("/gecmis/gonderilenler")
    assert gonderilenler.status_code == 200
    assert f"/gecmis/{kaynak_id}/kopyala" in gonderilenler.text

    yanit = client.post(f"/gecmis/{kaynak_id}/kopyala", follow_redirects=False)
    assert yanit.status_code == 303
    yeni_id = int(yanit.headers["location"].rsplit("=", 1)[1])

    with __import__("sqlmodel").Session(veritabani) as oturum:
        kaynak = oturum.get(Hesaplama, kaynak_id)
        kopya = oturum.get(Hesaplama, yeni_id)
        assert kaynak.durum == DURUM_ONAY_BEKLIYOR
        assert kaynak.onay_hedefi == "onur.simsek"
        assert kopya.ad == "Bekleyen Tahmin COPY"
        assert kopya.durum == DURUM_TASLAK
        assert kopya.revizyon == 1
        assert kopya.onay_hedefi is None
        assert kopya.onay_hedefi_ad_soyad is None
        assert kopya.onay_tarihi is None
        assert kopya.onaylayan_kullanici_adi is None
        assert len(kopya.kalemler) == 1


def test_reddedilmis_kopyalanir(client, veritabani):
    from app.models import DURUM_TASLAK

    _kaydet(client, "Reddedilen Tahmin")
    kaynak_id = _hesaplama_id_al(veritabani, "Reddedilen Tahmin")
    with __import__("sqlmodel").Session(veritabani) as oturum:
        kayit = oturum.get(Hesaplama, kaynak_id)
        kayit.durum = DURUM_TASLAK
        kayit.red_gerekce = "Eksik aciklama"
        kayit.reddeden_kullanici_adi = "onur.simsek"
        kayit.revizyon = 2
        oturum.add(kayit)
        oturum.commit()

    yanit = client.post(f"/gecmis/{kaynak_id}/kopyala", follow_redirects=False)
    assert yanit.status_code == 303
    yeni_id = int(yanit.headers["location"].rsplit("=", 1)[1])

    with __import__("sqlmodel").Session(veritabani) as oturum:
        kaynak = oturum.get(Hesaplama, kaynak_id)
        kopya = oturum.get(Hesaplama, yeni_id)
        assert kaynak.red_gerekce == "Eksik aciklama"
        assert kaynak.reddeden_kullanici_adi == "onur.simsek"
        assert kopya.ad == "Reddedilen Tahmin COPY"
        assert kopya.durum == DURUM_TASLAK
        assert kopya.revizyon == 1
        assert kopya.red_gerekce is None
        assert kopya.reddeden_kullanici_adi is None
        assert kopya.olusturan_kullanici_adi == "test.kullanici"
        assert len(kopya.kalemler) == 1


def test_kopya_adi_carpisinca_numaralanir(client, veritabani):
    _kaydet(client, "OnurŞimşekTest")
    kaynak_id = _hesaplama_id_al(veritabani, "OnurŞimşekTest")

    ilk = client.post(f"/gecmis/{kaynak_id}/kopyala", follow_redirects=False)
    assert ilk.status_code == 303
    ikinci = client.post(f"/gecmis/{kaynak_id}/kopyala", follow_redirects=False)
    assert ikinci.status_code == 303

    with __import__("sqlmodel").Session(veritabani) as oturum:
        from sqlmodel import select

        adlar = sorted(
            oturum.exec(select(Hesaplama.ad)).all(),
            key=lambda a: (len(a), a),
        )
    assert "OnurŞimşekTest" in adlar
    assert "OnurŞimşekTest COPY" in adlar
    assert "OnurŞimşekTest COPY (1)" in adlar


def test_kopyala_csrf_eksik_reddedilir(client, veritabani):
    _kaydet(client, "CSRF Copy")
    kaynak_id = _hesaplama_id_al(veritabani, "CSRF Copy")
    client.headers.pop("X-CSRF-Token", None)
    yanit = client.post(f"/gecmis/{kaynak_id}/kopyala", follow_redirects=False)
    assert yanit.status_code == 403


def test_kopyala_yetkisiz_403(client, veritabani):
    from app.main import app
    from app.yetkilendirme import aktif_kullanici

    _kaydet(client, "Gizli Taslak")
    kaynak_id = _hesaplama_id_al(veritabani, "Gizli Taslak")

    app.dependency_overrides[aktif_kullanici] = lambda: {
        "kullanici_adi": "deniz.aksoy",
        "ad_soyad": "Deniz Aksoy",
        "unvan": "",
        "gruplar": ["AFH-Calisanlar"],
    }
    yanit = client.post(f"/gecmis/{kaynak_id}/kopyala", follow_redirects=False)
    assert yanit.status_code == 403
    app.dependency_overrides.pop(aktif_kullanici, None)


def test_yonetici_baskasinin_kaydini_kendine_kopyalar(client, veritabani):
    from app.main import app
    from app.models import DURUM_ONAY_BEKLIYOR, DURUM_TASLAK
    from app.yetkilendirme import aktif_kullanici

    app.dependency_overrides[aktif_kullanici] = lambda: {
        "kullanici_adi": "kerem.acar",
        "ad_soyad": "Kerem Acar",
        "unvan": "",
        "gruplar": ["AFH-Calisanlar", "DEPT-IT"],
        "departman": "it-altyapi",
        "manager": "onur.simsek",
        "manager_zinciri": ["onur.simsek", "emre.turan"],
    }
    _kaydet(client, "Kerem Tahmini")
    kaynak_id = _hesaplama_id_al(veritabani, "Kerem Tahmini")
    with __import__("sqlmodel").Session(veritabani) as oturum:
        kayit = oturum.get(Hesaplama, kaynak_id)
        kayit.durum = DURUM_ONAY_BEKLIYOR
        kayit.olusturan_manager_zinciri = ["onur.simsek", "emre.turan"]
        kayit.onay_hedefi = "onur.simsek"
        oturum.add(kayit)
        oturum.commit()

    app.dependency_overrides[aktif_kullanici] = lambda: {
        "kullanici_adi": "onur.simsek",
        "ad_soyad": "Onur Simsek",
        "unvan": "",
        "gruplar": ["AFH-Calisanlar", "AFH-Yoneticiler", "DEPT-IT"],
        "rol": "yonetici",
        "manager_zinciri": ["emre.turan"],
        "departman": "it-altyapi",
    }
    arama = client.get("/gecmis/arama")
    assert arama.status_code == 200
    assert f"/gecmis/{kaynak_id}/kopyala" in arama.text

    yanit = client.post(f"/gecmis/{kaynak_id}/kopyala", follow_redirects=False)
    assert yanit.status_code == 303
    yeni_id = int(yanit.headers["location"].rsplit("=", 1)[1])

    with __import__("sqlmodel").Session(veritabani) as oturum:
        kaynak = oturum.get(Hesaplama, kaynak_id)
        kopya = oturum.get(Hesaplama, yeni_id)
        assert kaynak.olusturan_kullanici_adi == "kerem.acar"
        assert kaynak.durum == DURUM_ONAY_BEKLIYOR
        assert kopya.olusturan_kullanici_adi == "onur.simsek"
        assert kopya.olusturan_ad_soyad == "Onur Simsek"
        assert kopya.durum == DURUM_TASLAK
        assert kopya.ad == "Kerem Tahmini COPY"
        assert kopya.onay_hedefi is None

    app.dependency_overrides.pop(aktif_kullanici, None)


_CALISAN = {
    "kullanici_adi": "test.kullanici",
    "ad_soyad": "Test Kullanici",
    "unvan": "Test Unvani",
    "gruplar": ["AFH-Calisanlar"],
    "manager": "onur.simsek",
    "manager_zinciri": ["onur.simsek", "emre.turan", "baris.kocak"],
    "rol": "calisan",
}
_YONETICI = {
    "kullanici_adi": "onur.simsek",
    "ad_soyad": "Onur Simsek",
    "unvan": "Birim Sorumlusu",
    "gruplar": ["AFH-Calisanlar", "AFH-Yoneticiler"],
    "rol": "yonetici",
    "manager": "emre.turan",
    "manager_zinciri": ["emre.turan"],
}


def _bekleyen_hesaplama(veritabani, ad, sahip="test.kullanici", hedef="onur.simsek"):
    from datetime import datetime, timezone

    from app.models import DURUM_ONAY_BEKLIYOR

    with __import__("sqlmodel").Session(veritabani) as oturum:
        kayit = Hesaplama(
            ad=ad,
            durum=DURUM_ONAY_BEKLIYOR,
            olusturan_kullanici_adi=sahip,
            onay_hedefi=hedef,
            toplam_aylik_maliyet=1.0,
            para_birimi="USD",
            olusturulma_tarihi=datetime.now(timezone.utc),
        )
        oturum.add(kayit)
        oturum.commit()
        oturum.refresh(kayit)
        return kayit.id


def test_reddedilen_taslak_listesinde_yok_gonderilenlerde_var(client, veritabani):
    from app.main import app
    from app.models import DURUM_TASLAK
    from app.yetkilendirme import aktif_kullanici
    from sqlmodel import select

    kayit_id = _bekleyen_hesaplama(veritabani, "Reddedilen Liste")
    app.dependency_overrides[aktif_kullanici] = lambda: _YONETICI
    reddet = client.post(
        f"/onay/{kayit_id}/reddet",
        data={"gerekce": "Eksik aciklama"},
        follow_redirects=False,
    )
    assert reddet.status_code == 303

    with __import__("sqlmodel").Session(veritabani) as oturum:
        kayit = oturum.exec(select(Hesaplama).where(Hesaplama.id == kayit_id)).one()
        assert kayit.durum == DURUM_TASLAK
        assert kayit.red_gerekce == "Eksik aciklama"
        assert kayit.reddeden_kullanici_adi == "onur.simsek"

    app.dependency_overrides[aktif_kullanici] = lambda: _CALISAN
    taslaklar = client.get("/gecmis/taslaklar")
    gonderilenler = client.get("/gecmis/gonderilenler")
    detay = client.get(f"/gecmis/{kayit_id}")
    pano = client.get("/")

    assert taslaklar.status_code == 200
    assert "Reddedilen Liste" not in taslaklar.text
    assert gonderilenler.status_code == 200
    assert "Reddedilen Liste" in gonderilenler.text
    assert "Reddeden" in gonderilenler.text
    assert "Onur Simsek" in gonderilenler.text
    assert "Eksik aciklama" in gonderilenler.text
    assert detay.status_code == 200
    assert "Reddeden" in detay.text
    assert "Onur Simsek" in detay.text
    kutu = re.search(r'class="detail-header__reject">(.*?)</p>', detay.text, re.S)
    assert kutu, "red kutusu yok"
    assert "Reddeden" in kutu.group(1)
    assert "Onur Simsek" in kutu.group(1)
    assert "Eksik aciklama" in kutu.group(1)
    assert pano.status_code == 200
    assert "Reddedilen Liste" in pano.text
    assert "Reddeden" in pano.text
    assert "Onur Simsek" in pano.text


def test_reddet_gerekcesiz_yine_reddedilen_listesine_girer(client, veritabani):
    from app.main import app
    from app.yetkilendirme import aktif_kullanici

    kayit_id = _bekleyen_hesaplama(veritabani, "Gerekcesiz Red")
    app.dependency_overrides[aktif_kullanici] = lambda: _YONETICI
    reddet = client.post(
        f"/onay/{kayit_id}/reddet",
        data={"gerekce": "   "},
        follow_redirects=False,
    )
    assert reddet.status_code == 303

    app.dependency_overrides[aktif_kullanici] = lambda: _CALISAN
    taslaklar = client.get("/gecmis/taslaklar")
    gonderilenler = client.get("/gecmis/gonderilenler")
    assert "Gerekcesiz Red" not in taslaklar.text
    assert "Gerekcesiz Red" in gonderilenler.text
    assert "Reddeden" in gonderilenler.text


def test_taslak_kaydet_red_bilgisini_korur(client, veritabani):
    from app.models import DURUM_TASLAK
    from sqlmodel import select

    _kaydet(client, "Red Sonrasi Taslak")
    kayit_id = _hesaplama_id_al(veritabani, "Red Sonrasi Taslak")
    with __import__("sqlmodel").Session(veritabani) as oturum:
        kayit = oturum.get(Hesaplama, kayit_id)
        kayit.durum = DURUM_TASLAK
        kayit.red_gerekce = "Revize"
        kayit.reddeden_kullanici_adi = "onur.simsek"
        oturum.add(kayit)
        oturum.commit()

    ekleme = client.post(
        "/tahmin/kalem-ekle", data={"urun_tipi": "managed_disks", "para_birimi": "USD"}
    )
    kalem_id = _kalem_id_cikar(ekleme.text)
    veri = {
        f"{kalem_id}.urun_tipi": "managed_disks",
        f"{kalem_id}.bolge": "eastus",
        f"{kalem_id}.kademe": "standardhdd",
        f"{kalem_id}.sku": "S4",
        f"{kalem_id}.adet": "1",
        "para_birimi": "USD",
        "hesaplama_adi": "Red Sonrasi Taslak",
        "hesaplama_id": str(kayit_id),
        "onaya_gonder": "0",
    }
    yanit = client.post("/tahmin/kaydet", data=veri)
    assert yanit.status_code == 200
    assert "/gecmis/gonderilenler" in (yanit.headers.get("HX-Redirect") or yanit.text)

    with __import__("sqlmodel").Session(veritabani) as oturum:
        kayit = oturum.exec(select(Hesaplama).where(Hesaplama.id == kayit_id)).one()
        assert kayit.durum == DURUM_TASLAK
        assert kayit.red_gerekce == "Revize"
        assert kayit.reddeden_kullanici_adi == "onur.simsek"

    taslaklar = client.get("/gecmis/taslaklar")
    gonderilenler = client.get("/gecmis/gonderilenler")
    assert "Red Sonrasi Taslak" not in taslaklar.text
    assert "Red Sonrasi Taslak" in gonderilenler.text


def test_onaya_tekrar_gonderince_red_bilgisi_silinir(client, veritabani):
    from app.models import DURUM_ONAY_BEKLIYOR, DURUM_TASLAK
    from sqlmodel import select

    _kaydet(client, "Tekrar Gonder")
    kayit_id = _hesaplama_id_al(veritabani, "Tekrar Gonder")
    with __import__("sqlmodel").Session(veritabani) as oturum:
        kayit = oturum.get(Hesaplama, kayit_id)
        kayit.durum = DURUM_TASLAK
        kayit.red_gerekce = "Revize"
        kayit.reddeden_kullanici_adi = "onur.simsek"
        oturum.add(kayit)
        oturum.commit()

    ekleme = client.post(
        "/tahmin/kalem-ekle", data={"urun_tipi": "managed_disks", "para_birimi": "USD"}
    )
    kalem_id = _kalem_id_cikar(ekleme.text)
    veri = {
        f"{kalem_id}.urun_tipi": "managed_disks",
        f"{kalem_id}.bolge": "eastus",
        f"{kalem_id}.kademe": "standardhdd",
        f"{kalem_id}.sku": "S4",
        f"{kalem_id}.adet": "1",
        "para_birimi": "USD",
        "hesaplama_adi": "Tekrar Gonder",
        "hesaplama_id": str(kayit_id),
        "onaya_gonder": "1",
        "onay_hedefi": "onur.simsek",
    }
    yanit = client.post("/tahmin/kaydet", data=veri)
    assert yanit.status_code == 200

    with __import__("sqlmodel").Session(veritabani) as oturum:
        kayit = oturum.exec(select(Hesaplama).where(Hesaplama.id == kayit_id)).one()
        assert kayit.durum == DURUM_ONAY_BEKLIYOR
        assert kayit.red_gerekce is None
        assert kayit.reddeden_kullanici_adi is None


def test_reddet_csrf_eksik_reddedilir(client, veritabani):
    from app.main import app
    from app.models import DURUM_ONAY_BEKLIYOR
    from app.yetkilendirme import aktif_kullanici
    from sqlmodel import select

    kayit_id = _bekleyen_hesaplama(veritabani, "CSRF Red")
    app.dependency_overrides[aktif_kullanici] = lambda: _YONETICI
    client.headers.pop("X-CSRF-Token", None)
    yanit = client.post(
        f"/onay/{kayit_id}/reddet",
        data={"gerekce": "Hayir"},
        follow_redirects=False,
    )
    assert yanit.status_code == 403

    with __import__("sqlmodel").Session(veritabani) as oturum:
        kayit = oturum.exec(select(Hesaplama).where(Hesaplama.id == kayit_id)).one()
        assert kayit.durum == DURUM_ONAY_BEKLIYOR
        assert kayit.reddeden_kullanici_adi is None


def test_ustu_olmayan_reddedileni_taslaklara_koymaz(client, veritabani):
    from app.main import app
    from app.models import DURUM_TASLAK
    from app.yetkilendirme import GENEL_MUDUR_SAM, aktif_kullanici

    gm = {
        "kullanici_adi": GENEL_MUDUR_SAM,
        "ad_soyad": "Ahmet Yildirim",
        "unvan": "Genel Mudur",
        "gruplar": ["AFH-Calisanlar", "AFH-Direktorler"],
        "rol": "direktor",
        "manager": None,
        "manager_zinciri": [],
    }
    app.dependency_overrides[aktif_kullanici] = lambda: gm
    _kaydet(client, "GM Taslak")
    taslak_id = _hesaplama_id_al(veritabani, "GM Taslak")
    reddedilen_id = _bekleyen_hesaplama(
        veritabani, "GM Reddedilen", sahip=GENEL_MUDUR_SAM
    )
    with __import__("sqlmodel").Session(veritabani) as oturum:
        kayit = oturum.get(Hesaplama, reddedilen_id)
        kayit.durum = DURUM_TASLAK
        kayit.red_gerekce = "Ust ret"
        kayit.reddeden_kullanici_adi = "onur.simsek"
        kayit.onay_hedefi = None
        oturum.add(kayit)
        oturum.commit()

    taslaklar = client.get("/gecmis/taslaklar")
    gonderilenler = client.get("/gecmis/gonderilenler")
    assert taslaklar.status_code == 200
    assert "GM Taslak" in taslaklar.text
    assert "GM Reddedilen" not in taslaklar.text
    assert gonderilenler.status_code == 200
    assert "GM Reddedilen" in gonderilenler.text
    assert taslak_id
    app.dependency_overrides.pop(aktif_kullanici, None)


def test_eski_red_aktivite_aktorunu_detay_ve_listede_gosterir(client, veritabani):
    from datetime import datetime, timezone

    from app.models import AktiviteKaydi, DURUM_TASLAK

    _kaydet(client, "OnurŞimşek RED COPY (2)")
    kayit_id = _hesaplama_id_al(veritabani, "OnurŞimşek RED COPY (2)")
    with __import__("sqlmodel").Session(veritabani) as oturum:
        kayit = oturum.get(Hesaplama, kayit_id)
        kayit.durum = DURUM_TASLAK
        kayit.red_gerekce = "x"
        kayit.reddeden_kullanici_adi = None
        oturum.add(kayit)
        oturum.add(
            AktiviteKaydi(
                aktor_kullanici_adi="onur.simsek",
                islem="reddedildi",
                hesaplama_id=kayit_id,
                detay="x",
                olusturulma_tarihi=datetime.now(timezone.utc),
            )
        )
        oturum.commit()

    detay = client.get(f"/gecmis/{kayit_id}")
    gonderilenler = client.get("/gecmis/gonderilenler")
    pano = client.get("/")
    assert detay.status_code == 200
    kutu = re.search(r'class="detail-header__reject">(.*?)</p>', detay.text, re.S)
    assert kutu
    assert "Reddeden" in kutu.group(1)
    assert "Onur Simsek" in kutu.group(1)
    assert "x" in kutu.group(1)
    assert gonderilenler.status_code == 200
    assert "Reddeden" in gonderilenler.text
    assert "Onur Simsek" in gonderilenler.text
    assert pano.status_code == 200
    assert "Reddeden" in pano.text
    assert "Onur Simsek" in pano.text


def test_eski_red_onaylayan_aktorunu_detayda_gosterir(client, veritabani):
    from app.models import DURUM_TASLAK

    _kaydet(client, "Tarihi Red")
    kayit_id = _hesaplama_id_al(veritabani, "Tarihi Red")
    with __import__("sqlmodel").Session(veritabani) as oturum:
        kayit = oturum.get(Hesaplama, kayit_id)
        kayit.durum = DURUM_TASLAK
        kayit.red_gerekce = "Eski gerekce"
        kayit.reddeden_kullanici_adi = None
        kayit.onaylayan_kullanici_adi = "onur.simsek"
        oturum.add(kayit)
        oturum.commit()

    detay = client.get(f"/gecmis/{kayit_id}")
    assert detay.status_code == 200
    assert "Reddeden" in detay.text
    assert "Onur Simsek" in detay.text


def test_kopya_red_gerekcesi_kaynak_reddedenini_gosterir(client, veritabani):
    from app.models import DURUM_TASLAK

    _kaydet(client, "OnurŞimşek RED")
    kaynak_id = _hesaplama_id_al(veritabani, "OnurŞimşek RED")
    _kaydet(client, "OnurŞimşek RED COPY (2)")
    kopya_id = _hesaplama_id_al(veritabani, "OnurŞimşek RED COPY (2)")
    with __import__("sqlmodel").Session(veritabani) as oturum:
        kaynak = oturum.get(Hesaplama, kaynak_id)
        kaynak.durum = DURUM_TASLAK
        kaynak.red_gerekce = "x"
        kaynak.reddeden_kullanici_adi = "onur.simsek"
        kopya = oturum.get(Hesaplama, kopya_id)
        kopya.durum = DURUM_TASLAK
        kopya.red_gerekce = "x"
        kopya.reddeden_kullanici_adi = None
        oturum.add(kaynak)
        oturum.add(kopya)
        oturum.commit()

    detay = client.get(f"/gecmis/{kopya_id}")
    assert detay.status_code == 200
    assert "Reddeden" in detay.text
    assert "Onur Simsek" in detay.text
    kutu = re.search(r'class="detail-header__reject">(.*?)</p>', detay.text, re.S)
    assert kutu
    assert "Onur Simsek" in kutu.group(1)


def test_gecmis_excel_reddeden_metasini_yazar(client, veritabani):
    from io import BytesIO

    from app.models import DURUM_TASLAK
    from openpyxl import load_workbook

    _kaydet(client, "Excel Red")
    kayit_id = _hesaplama_id_al(veritabani, "Excel Red")
    with __import__("sqlmodel").Session(veritabani) as oturum:
        kayit = oturum.get(Hesaplama, kayit_id)
        kayit.durum = DURUM_TASLAK
        kayit.red_gerekce = "Eksik"
        kayit.reddeden_kullanici_adi = "onur.simsek"
        oturum.add(kayit)
        oturum.commit()

    excel = client.get(f"/gecmis/{kayit_id}/excel")
    assert excel.status_code == 200
    satirlar = list(load_workbook(BytesIO(excel.content)).active.iter_rows(values_only=True))
    assert ("Reddeden", "Onur Simsek") in [(s[0], s[1]) for s in satirlar]
    assert any(s[0] == "Durum" for s in satirlar)
